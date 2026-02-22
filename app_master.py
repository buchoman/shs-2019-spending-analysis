"""
Survey of Household Spending - Master App (2019 / 2021)
This application allows users to select survey year and demographic attributes to get
detailed estimates of average household spending with bootstrap variance estimates.
"""

import streamlit as st
import pandas as pd
import numpy as np
from pathlib import Path
import os
from io import BytesIO
import json
import warnings
warnings.filterwarnings('ignore')

import pyreadstat
from config.year_config_2019 import YearConfig2019
from config.year_config_2021 import YearConfig2021
from loaders.fixed_width_loader import load_fixed_width_pumf, load_fixed_width_bsw

# Set page config
st.set_page_config(
    page_title="Survey of Household Spending - Spending Estimates",
    page_icon="💰",
    layout="wide"
)

# --- Version selector (mandatory, must select before any other UI) ---
st.header("Select Survey Year")
year_choice = st.radio(
    "Survey Version",
    options=[2019, 2021],
    format_func=lambda x: str(x),
    index=None,
    key="year_selector",
    help="Select the survey year to analyze. This selection is required before proceeding.",
)
if year_choice is None:
    st.warning("Please select a survey year (2019 or 2021) to continue.")
    st.stop()

# Year configuration based on selection
YEAR_CONFIG = YearConfig2019() if year_choice == 2019 else YearConfig2021()

# Data paths from config
_paths = YEAR_CONFIG.get_data_paths()
DATA_DIR = _paths["data_dir"]
MAIN_FILE = _paths["main_file"]
BSW_FILE = _paths["bsw_file"]
LAYOUT_FILE = _paths.get("layout_file")
BSW_LAYOUT_FILE = _paths.get("bsw_layout_file")

# Allocation Input: cache file from config
ALLOCATION_INPUT_CACHE = Path(YEAR_CONFIG.get_allocation_cache_filename())

# Value label mappings for filter variables (from year config)
VALUE_LABELS = YEAR_CONFIG.get_value_labels()

def format_value(var_name, value):
    """Format a value using its label if available. Handles 2021 data where codes may be '01', 1, or 1.0."""
    # Map canonical column names to value label keys (config uses e.g. RP_AGEGRP, app uses RP_AgeGrp)
    _canonical_to_key = {
        "RP_AgeGrp": "RP_AGEGRP", "RP_Gender": "RP_GENDER", "RP_MarStat": "RP_MARSTAT", "RP_Educ": "RP_EDUC",
        "SP_AgeGrp": "SP_AGEGRP", "SP_Gender": "SP_GENDER", "SP_Educ": "SP_EDUC",
        "Prov": "PROV", "HHType6": "HHTYPE6", "HHSize": "HHSIZE", "DwellTyp": "DWELTYP", "Tenure": "TENURE",
        "Numbedr": "NUMBEDR", "VehicleYN": "VEHICLEYN", "RecVehYN": "RECVEHYN",
        "HH_MajIncSrc": "HH_MAJINCSRC", "P0to4YN": "P0TO4YN", "P5to15YN": "P5TO15YN",
    }
    key = var_name if var_name in VALUE_LABELS else _canonical_to_key.get(var_name, var_name)
    if key not in VALUE_LABELS:
        return str(value)
    labels = VALUE_LABELS[key]
    # Normalize value: handle numpy int/float, 1.0 -> "1", etc.
    try:
        v = float(value)
        if v == int(v):
            str_val = str(int(v))
        else:
            str_val = str(value).strip()
    except (TypeError, ValueError):
        str_val = str(value).strip()
    if str_val in labels:
        return labels[str_val]
    # Try normalized forms: "1" <-> "01", "2" <-> "02" for 2-digit codes
    if str_val.replace(".", "").isdigit() or str_val.isdigit():
        cand = str_val.zfill(2) if len(str_val) <= 2 else str_val
        if cand in labels:
            return labels[cand]
        if str_val.zfill(2) in labels:
            return labels[str_val.zfill(2)]
    return str(value)

# Parent/aggregate variables that should be excluded from category totals (to avoid double-counting)
# These are totals that include their subcategories
PARENT_TOTALS = {
    "FD001",  # Food expenditures (parent of all FD variables)
    "FD003",  # Food purchased from stores (parent of store food items)
    "FD990",  # Food purchased from restaurants (parent of FD991)
    "FD991",  # Restaurant meals (parent of FD992-FD995)
    "FD100",  # Bakery products (parent of FD101-FD108, FD112)
    "FD200",  # Cereal grains and cereal products (parent of FD201-FD212)
    "FD300",  # Fruit, fruit preparations and nuts (parent of FD301-FD316, FD330-FD382)
    "FD400",  # Vegetables and vegetable preparations (parent of FD401-FD412, FD418, FD421, FD440-FD479)
    "FD500",  # Dairy products and eggs (parent of FD501-FD505, FD520-FD525, FD540-FD541, FD550-FD555, FD570-FD572)
    "FD600",  # Meat (parent of FD601-FD607, FD650-FD660)
    "FD700",  # Fish and seafood (parent of FD701-FD706, FD720-FD724, FD730-FD732)
    "FD800",  # Non-alcoholic beverages and other food products (parent of FD801-FD802, FD806, FD814-FD815, FD821, FD827-FD829, FD833-FD889)
    "CS030",  # Communications (parent of CS003, CS004, CS005, etc.)
    "HF001",  # Household furnishings and equipment (parent of HF002)
    "HE001",  # Household equipment (parent of HE002, HE010, etc.)
    "HC001",  # Health care (parent of HC002, HC022, etc.)
    "HO001",  # Household operations (parent of HO002, HO003, etc.)
    "PC001",  # Personal care (parent of PC002, PC020)
    "RE001",  # Recreation (parent of RE002, RE003, etc.)
    "RO001",  # Reading materials (parent of RO002, RO003, etc.)
    "RV001",  # Recreational vehicles (parent of RV010, RV020)
    "SH001",  # Shelter (parent of SH002, SH003, etc.)
    "SH002",  # Shelter (another parent total)
    "TR001",  # Transportation (parent of TR002, TR003, etc.)
    "TR002",  # Transportation (another parent total)
    "ME001",  # Miscellaneous expenditures (parent of ME039, ME040)
    "TA018",  # Tobacco products, alcoholic beverages and cannabis (parent of TA005, TA006, TA007, TA008, TA990)
    # TE001 and TC001 removed: user-requested top-level totals (Total expenditure, Total current consumption) to be shown
}

# Spending category mappings (organized by major category prefix)
# Excludes parent totals to avoid double-counting (except Totals: TE001, TC001 which are shown by request)
SPENDING_CATEGORIES = {
    "Totals": ["TE001", "TC001"],  # Total expenditure, Total current consumption
    "Child Care": ["CC001"],
    "Clothing": ["CL014", "CL015", "CL016", "CL017", "CL023", "CL026", "CL029", "CL030", "CL990"],
    "Communications": ["CS003", "CS004", "CS005", "CS007", "CS008", "CS020", "CS021"],  # Excluded CS030 (parent)
    "Education": ["ED002", "ED003", "ED030"],
    "Personal Insurance": ["EP011"],
    "Food": [
        # Excluded parent totals: FD001, FD003, FD990, FD991, FD100, FD200, FD300, FD400, FD500, FD600, FD700, FD800
        "FD1001", "FD1002", "FD1003", "FD1004", "FD101", "FD102", "FD103",
        "FD104", "FD105", "FD106", "FD107", "FD108", "FD112", "FD201", "FD202", "FD203",
        "FD204", "FD205", "FD206", "FD207", "FD208", "FD209", "FD212", "FD301", "FD302",
        "FD303", "FD304", "FD305", "FD308", "FD309", "FD315", "FD316", "FD330", "FD331", "FD350",
        "FD380", "FD381", "FD382", "FD401", "FD402", "FD403", "FD404", "FD405", "FD406",
        "FD407", "FD408", "FD409", "FD410", "FD411", "FD412", "FD418", "FD421", "FD440", "FD441",
        "FD442", "FD447", "FD470", "FD471", "FD478", "FD479", "FD501", "FD502", "FD503",
        "FD504", "FD505", "FD520", "FD521", "FD522", "FD525", "FD540", "FD541", "FD550", "FD551",
        "FD555", "FD570", "FD571", "FD572", "FD601", "FD602", "FD603", "FD604", "FD607",
        "FD650", "FD651", "FD660", "FD701", "FD705", "FD706", "FD720", "FD721", "FD722",
        "FD723", "FD724", "FD730", "FD731", "FD732", "FD801", "FD802", "FD806", "FD814",
        "FD815", "FD821", "FD827", "FD828", "FD829", "FD833", "FD834", "FD835", "FD836", "FD837",
        "FD838", "FD839", "FD840", "FD841", "FD842", "FD843", "FD844", "FD845", "FD846", "FD847",
        "FD850", "FD851", "FD852", "FD853", "FD854", "FD855", "FD857", "FD870", "FD871", "FD872",
        "FD873", "FD874", "FD875", "FD879", "FD880", "FD881", "FD882", "FD883", "FD884", "FD885",
        "FD889", "FD992", "FD993", "FD994", "FD995"  # Restaurant subcategories (excluded FD990, FD991)
    ],
    "Games of Chance": ["GC001"],
    "Health Care": ["HC002", "HC022", "HC025", "HC061"],  # Excluded HC001 (parent total)
    "Household Equipment": ["HE002", "HE010", "HE017", "HE020"],  # Excluded HE001 (parent total)
    "Household Furnishings": ["HF002"],  # Excluded HF001 (parent total)
    "Household Operations": ["HO002", "HO003", "HO004", "HO005", "HO006", "HO010", "HO014", "HO018", "HO022"],  # Excluded HO001 (parent total)
    "Miscellaneous": ["ME039", "ME040"],  # Excluded ME001 (parent total)
    "Gifts and Contributions": ["MG001"],
    "Personal Care": ["PC002", "PC020"],  # Excluded PC001 (parent total)
    "Recreation": [
        # Excluded RE001 (parent total)
        "RE002", "RE003", "RE006", "RE007", "RE010", "RE016", "RE020", "RE022", "RE032",
        "RE040", "RE041", "RE052", "RE060", "RE061", "RE062", "RE063", "RE066", "RE067", "RE074",
        "RE090", "RE120", "RE124", "RE127", "RE140", "RE990"
    ],
    "Reading Materials": ["RO002", "RO003", "RO004", "RO005", "RO010"],  # Excluded RO001 (parent total)
    "Recreational Vehicles": ["RV010", "RV020"],  # Excluded RV001 (parent total)
    "Shelter": [
        # Excluded SH001, SH002 (parent totals)
        "SH003", "SH004", "SH010", "SH011", "SH015", "SH016", "SH019", "SH030",
        "SH031", "SH032", "SH033", "SH034", "SH040", "SH041", "SH042", "SH044", "SH046", "SH047",
        "SH050", "SH060", "SH061", "SH062", "SH082", "SH990", "SH991", "SH992"
    ],
    "Tobacco and Alcohol": ["TA005", "TA006", "TA007", "TA008", "TA990"],  # Excluded TA018 (parent total)
    "Transportation": [
        # Excluded TR001, TR002 (parent totals)
        "TR003", "TR004", "TR008", "TR010", "TR020", "TR021", "TR022", "TR030",
        "TR031", "TR033", "TR034", "TR036", "TR038", "TR039", "TR070", "TR071", "TR085"
    ],
    "Income Taxes": ["TX010"]
}

# Spending variable descriptions (from SAS labels)
SPENDING_DESCRIPTIONS = {
    "CC001": "Child care",
    "CL014": "Laundromats, dry-cleaning and laundry services",
    "CL015": "Services for clothing, footwear and jewellery",
    "CL016": "Clothing services",
    "CL017": "Clothing material, yarn, thread and other notions",
    "CL023": "Children's wear (under 14 years)",
    "CL026": "Men's and boys' wear (14 years and over)",
    "CL029": "Women's and girls' wear (14 years and over)",
    "CL030": "Clothing and accessories",
    "CL990": "Accessories, watches, jewellery and athletic footwear",
    "CS003": "Telephone",
    "CS004": "Landline telephone services",
    "CS005": "Cell phone and pager services",
    "CS007": "Internet access services",
    "CS008": "Digital services",
    "CS020": "Postal, courier, delivery and other communication services",
    "CS021": "Telephones and equipment",
    "CS030": "Communications",
    "ED002": "Education",
    "ED003": "Tuition fees",
    "ED030": "Textbooks and school supplies",
    "EP011": "Personal insurance payments and pension contributions",
    "FD001": "Food expenditures",
    "FD003": "Food purchased from stores",
    "FD100": "Bakery products",
    "FD1001": "Frozen side dishes and other frozen prepared food",
    "FD1002": "Other ready-to-serve prepared food",
    "FD1003": "Cod, flounder, sole and haddock (fresh or frozen, uncooked)",
    "FD1004": "Other oils and fats",
    "FD101": "Bread and unsweetened rolls and buns",
    "FD102": "Bread",
    "FD103": "Unsweetened rolls and buns",
    "FD104": "Cookies and crackers",
    "FD105": "Cookies and sweet biscuits",
    "FD106": "Crackers and crisp breads",
    "FD107": "Other bakery products",
    "FD108": "Other bakery products (except frozen)",
    "FD112": "Frozen bakery products",
    "FD200": "Cereal grains and cereal products",
    "FD201": "Rice and rice mixes",
    "FD202": "Rice",
    "FD203": "Rice mixes",
    "FD204": "Pasta products",
    "FD205": "Pasta (fresh or dry)",
    "FD206": "Pasta (canned)",
    "FD207": "Pasta mixes",
    "FD208": "Other cereal grains and cereal products",
    "FD209": "Flour and flour-based mixes",
    "FD212": "Breakfast cereal and other grain products (except infant)",
    "FD300": "Fruit, fruit preparations and nuts",
    "FD301": "Fresh fruit",
    "FD302": "Apples (fresh)",
    "FD303": "Bananas and plantains (fresh)",
    "FD304": "Grapes (fresh)",
    "FD305": "Peaches and nectarines (fresh)",
    "FD308": "Pears (fresh)",
    "FD309": "Berries (fresh)",
    "FD315": "Citrus fruit (fresh)",
    "FD316": "Other fruit (fresh)",
    "FD330": "Preserved fruit and fruit preparations",
    "FD331": "Fruit juice",
    "FD350": "Other preserved fruit and fruit preparations",
    "FD380": "Nuts and seeds",
    "FD381": "Peanuts (shelled or unshelled)",
    "FD382": "Other nuts and seeds",
    "FD400": "Vegetables and vegetable preparations",
    "FD401": "Fresh vegetables",
    "FD402": "Potatoes (except sweet potatoes)",
    "FD403": "Tomatoes (fresh)",
    "FD404": "Lettuce (fresh)",
    "FD405": "Cabbage (fresh)",
    "FD406": "Carrots (fresh)",
    "FD407": "Onions (fresh)",
    "FD408": "Celery (fresh)",
    "FD409": "Cucumber (fresh)",
    "FD410": "Mushrooms (fresh)",
    "FD411": "Broccoli (fresh)",
    "FD412": "Other vegetables (fresh)",
    "FD418": "Peppers (fresh)",
    "FD421": "Fresh herbs",
    "FD440": "Frozen and dried vegetables",
    "FD441": "Potato products (frozen)",
    "FD442": "Other frozen vegetables",
    "FD447": "Dried vegetables and legumes",
    "FD470": "Canned vegetables and other vegetable preparations",
    "FD471": "Canned or bottled vegetables",
    "FD478": "Ready-to-serve or ready-to-cook prepared salads and side dishes, fruit or vegetable based",
    "FD479": "Vegetable juice (canned or bottled)",
    "FD500": "Dairy products and eggs",
    "FD501": "Cheese",
    "FD502": "Cheddar cheese",
    "FD503": "Mozzarella cheese",
    "FD504": "Processed cheese",
    "FD505": "Other cheeses",
    "FD520": "Milk",
    "FD521": "Fluid whole milk",
    "FD522": "Fluid low-fat milk",
    "FD525": "Skim and other fluid milk",
    "FD540": "Butter",
    "FD541": "Ice cream and ice milk (including novelties)",
    "FD550": "Other dairy products",
    "FD551": "Other processed milk products",
    "FD555": "Other processed dairy products",
    "FD570": "Eggs and other egg products",
    "FD571": "Eggs",
    "FD572": "Other egg products",
    "FD600": "Meat",
    "FD601": "Meat (except processed meat)",
    "FD602": "Beef",
    "FD603": "Pork",
    "FD604": "Poultry",
    "FD607": "Other meat and poultry",
    "FD650": "Processed meat",
    "FD651": "Bacon and ham",
    "FD660": "Other processed meat",
    "FD700": "Fish and seafood",
    "FD701": "Fresh or frozen fish",
    "FD705": "Salmon (fresh or frozen, uncooked)",
    "FD706": "Other fish (fresh or frozen, uncooked)",
    "FD720": "Canned fish or other preserved fish",
    "FD721": "Tuna (canned)",
    "FD722": "Salmon (canned)",
    "FD723": "Other fish (canned or bottled)",
    "FD724": "Cured fish",
    "FD730": "Seafood and other marine products",
    "FD731": "Shrimp and prawns",
    "FD732": "Other seafood and marine products",
    "FD800": "Non-alcoholic beverages and other food products",
    "FD801": "Non-alcoholic beverages and beverage mixes",
    "FD802": "Coffee and tea",
    "FD806": "Non-alcoholic beverages",
    "FD814": "Sugar and confectionery",
    "FD815": "Sugar, syrups and sugar substitutes",
    "FD821": "Candies and chocolates",
    "FD827": "Margarine, oils and fats (excluding butter)",
    "FD828": "Margarine",
    "FD829": "Cooking and salad oils",
    "FD833": "Condiments, spices and vinegars",
    "FD834": "Mayonnaise, salad dressings and dips",
    "FD835": "Pasta and pizza sauces (canned, bottled or dried)",
    "FD836": "Other sauces and gravies (canned, bottled or dried)",
    "FD837": "Dried herbs and spices",
    "FD838": "Ketchup",
    "FD839": "Other condiments (including vinegar)",
    "FD840": "Pickled vegetables (including olives)",
    "FD841": "Infant food",
    "FD842": "Infant formula",
    "FD843": "Infant cereals and biscuits",
    "FD844": "Canned or bottled infant food",
    "FD845": "Frozen prepared food",
    "FD846": "Frozen dinners and entrees",
    "FD847": "Frozen pizza",
    "FD850": "Soup (except infant soup)",
    "FD851": "Soup (chilled, frozen, canned or bottled)",
    "FD852": "Soup (dried)",
    "FD853": "Ready-to-serve prepared food",
    "FD854": "Dinners and entrees (except frozen)",
    "FD855": "Pizza (except frozen)",
    "FD857": "Fish portions (pre-cooked and frozen)",
    "FD870": "Other food preparations",
    "FD871": "Peanut butter and other nut butters",
    "FD872": "Honey",
    "FD873": "Flavoured drink powders, crystals and syrups",
    "FD874": "Non-dairy frozen ice treats",
    "FD875": "Dessert powders",
    "FD879": "Food seasonings (including table salt)",
    "FD880": "Other materials for food preparation",
    "FD881": "Tofu",
    "FD882": "Other canned, bottled or dried meals",
    "FD883": "Snack food",
    "FD884": "Potato-based snack foods",
    "FD885": "Other snack foods",
    "FD889": "Other infant food (including frozen)",
    "FD990": "Food purchased from restaurants",
    "FD991": "Restaurant meals",
    "FD992": "Restaurant dinners",
    "FD993": "Restaurant lunches",
    "FD994": "Restaurant breakfasts",
    "FD995": "Restaurant snacks and beverages",
    "GC001": "Games of chance",
    "HC001": "Health care",
    "HC002": "Direct costs to household",
    "HC022": "Private health insurance plan premiums",
    "HC025": "Accident or disability insurance premiums",
    "HC061": "Private health and dental plan premiums",
    "HE001": "Household equipment",
    "HE002": "Household appliances",
    "HE010": "Other household equipment",
    "HE017": "Maintenance, rental, repairs and services related to household furnishings and equipment",
    "HE020": "Services related to household furnishings and equipment",
    "HF001": "Household furnishings and equipment",
    "HF002": "Household furnishings",
    "HO001": "Household operations",
    "HO002": "Domestic and other custodial services (excluding child care)",
    "HO003": "Pet expenses",
    "HO004": "Pet food",
    "HO005": "Purchase of pets and pet-related goods",
    "HO006": "Veterinarian and other services",
    "HO010": "Household cleaning supplies and equipment",
    "HO014": "Paper, plastic and foil supplies",
    "HO018": "Garden supplies and services",
    "HO022": "Other household supplies",
    "ME001": "Miscellaneous expenditures",
    "ME039": "Financial services",
    "ME040": "Other miscellaneous goods and services",
    "MG001": "Gifts of money, support payments and charitable contributions",
    "PC001": "Personal care",
    "PC002": "Personal care products",
    "PC020": "Personal care services",
    "RE001": "Recreation",
    "RE002": "Recreational equipment and related services",
    "RE003": "Sports, athletic and recreation equipment",
    "RE006": "Video game systems and accessories (excluding for computers)",
    "RE007": "Art and craft materials",
    "RE010": "Computer equipment and supplies",
    "RE016": "Photographic goods and services",
    "RE020": "Photographic services",
    "RE022": "Collectors' items (e.g. stamps, coins)",
    "RE032": "Other recreational equipment",
    "RE040": "Home entertainment equipment and services",
    "RE041": "Home entertainment equipment",
    "RE052": "Home entertainment services",
    "RE060": "Recreational services",
    "RE061": "Entertainment",
    "RE062": "Movie theatres",
    "RE063": "Live sporting and performing arts events",
    "RE066": "Admission fees to museums, zoos, and other sites",
    "RE067": "Television and satellite radio services (including installation, service and pay TV charges)",
    "RE074": "Package trips",
    "RE090": "Use of recreational facilities and fees for other recreational activities",
    "RE120": "Camcorders, cameras, parts, accessories and related equipment",
    "RE124": "Sports, athletic and recreational equipment and related services",
    "RE127": "Rental, maintenance and repairs of sports, athletic and recreational equipment",
    "RE140": "Other recreational services",
    "RE990": "Outdoor play equipment and children's toys",
    "RO001": "Reading materials and other printed matter",
    "RO002": "Newspapers",
    "RO003": "Magazines and periodicals",
    "RO004": "Books and E-Books (excluding school books)",
    "RO005": "Maps, sheet music and other printed matter",
    "RO010": "Services related to reading materials (e.g. photocopying, library fees)",
    "RV001": "Recreational vehicles and associated services",
    "RV010": "Operation of recreational vehicles",
    "RV020": "Purchase of recreational vehicles",
    "SH001": "Shelter",
    "SH002": "Principal accommodation",
    "SH003": "Rented living quarters",
    "SH004": "Rent",
    "SH010": "Owned living quarters",
    "SH011": "Mortgage paid",
    "SH015": "Homeowners' insurance premiums",
    "SH016": "Other expenditures for owned living quarters",
    "SH019": "Mortgage insurance premiums",
    "SH030": "Water, fuel and electricity for principal accommodation",
    "SH031": "Water and sewage",
    "SH032": "Electricity",
    "SH033": "Natural gas",
    "SH034": "Other fuel",
    "SH040": "Other accommodation",
    "SH041": "Owned secondary residences",
    "SH042": "Mortgage paid",
    "SH044": "Insurance premiums",
    "SH046": "Other expenses for owned secondary residences",
    "SH047": "Other owned properties",
    "SH050": "Accommodation away from home",
    "SH060": "Communication and home security services (e.g. landline telephone, television, satellite radio and Internet)",
    "SH061": "Property and school taxes, water and sewage charges",
    "SH062": "Electricity and fuel (e.g. natural gas and wood)",
    "SH082": "Repairs and maintenance",
    "SH990": "Other expenses for rented living quarters",
    "SH991": "Condominium fees, property taxes and school taxes",
    "SH992": "All other expenses for the owned living quarters",
    "TA005": "Alcoholic beverages",
    "TA006": "Alcoholic beverages served on licensed premises and in restaurants",
    "TA007": "Alcoholic beverages purchased from stores",
    "TA008": "Self-made alcoholic beverages",
    "TA018": "Tobacco products, alcoholic beverages and cannabis for non-medical use",
    "TA990": "Tobacco products, smokers' supplies and cannabis for non-medical use",
    "TR001": "Transportation",
    "TR002": "Private transportation",
    "TR003": "Private use automobiles, vans and trucks",
    "TR004": "Purchase of automobiles, vans and trucks",
    "TR008": "Accessories for automobiles, vans and trucks",
    "TR010": "Fees for leased automobiles, vans and trucks",
    "TR020": "Rented automobiles, vans and trucks",
    "TR021": "Fees for rented vehicles (including insurance and mileage)",
    "TR022": "Other expenses for rented automobiles, vans and trucks",
    "TR030": "Automobile, van and truck operations",
    "TR031": "Registration fees (including insurance if part of registration)",
    "TR033": "Tires, batteries, and other parts and supplies for vehicles",
    "TR034": "Maintenance and repairs of vehicles",
    "TR036": "Gas and other fuels (all vehicles and tools)",
    "TR038": "Parking (excluding parking fees included in rent and traffic and parking tickets)",
    "TR039": "Drivers' licences and tests, and driving lessons",
    "TR070": "Public transportation",
    "TR071": "Vehicle operation, security and communication services",
    "TR085": "Public and private vehicle insurance premiums",
    "TX010": "Income taxes",
    "TC001": "Total current consumption",
    "TE001": "Total expenditure"
}

# Get all spending variables (apply 2021 code mapping where variables differ)
def _map_spending_codes(codes_iter):
    """Apply year-specific spending code mapping (2019->2021 where different)."""
    mapping = YEAR_CONFIG.get_spending_code_mapping()
    result = []
    for c in codes_iter:
        mapped = mapping.get(c, c)
        if mapped is not None:
            result.append(mapped)
    return result

ALL_SPENDING_VARS = []
for category, vars_list in SPENDING_CATEGORIES.items():
    ALL_SPENDING_VARS.extend(_map_spending_codes(vars_list))
ALL_SPENDING_VARS = sorted(set(ALL_SPENDING_VARS))

# Items to include - only the 19 specified expenditure categories
# These are Level 2 categories plus totals that should balance with TC001
ITEMS_FOR_TC001_BALANCE = {
    # Totals
    "TE001",   # Total expenditure
    "TC001",   # Total current consumption
    
    # Level 2 expenditure categories
    "FD001",   # Food expenditures
    "SH001",   # Shelter
    "HO001",   # Household operations
    "HF001",   # Household furnishings and equipment
    "CL030",   # Clothing and accessories
    "TR001",   # Transportation
    "HC001",   # Health care
    "PC001",   # Personal care
    "RE001",   # Recreation
    "ED002",   # Education
    "RO001",   # Reading materials and other printed matter
    "TA018",   # Tobacco products, alcoholic beverages and cannabis for non-medical use
    "GC001",   # Games of chance
    "ME001",   # Miscellaneous expenditures
    "TX010",   # Income taxes
    "EP011",   # Personal insurance payments and pension contributions
    "MG001"    # Gifts of money, support payments and charitable contributions
}

# No parent totals to exclude - we're using Level 2 categories directly
PARENT_TOTALS_TO_EXCLUDE = set()


def _to_float(x):
    """Convert to float; return None if not parseable."""
    if x is None or (isinstance(x, float) and np.isnan(x)):
        return None
    try:
        return float(x)
    except (ValueError, TypeError):
        return None


def parse_allocation_input_excel(uploaded_file):
    """
    Parse an Allocation Input Form Excel file. Expects:
    - A column for expenditure category (Spending Code / Code / Variable / Category or first column)
    - A column for Shared Consumption % (header containing 'shared' and '%' or 'pct' or 'consumption')
    - A column for Child Intensity Index (header containing 'child' and 'intensity')
    Returns: {var_code: {'shared_pct': float|None, 'child_intensity': float|None}}
    """
    try:
        # Detect header row: first row containing "Spending Code" (form may have a title row above)
        df_peek = pd.read_excel(uploaded_file, sheet_name=0, header=None, nrows=15)
    except Exception as e:
        return None, str(e)
    header_row = 0
    for i in range(len(df_peek)):
        for v in df_peek.iloc[i]:
            if v is None:
                continue
            if isinstance(v, float) and np.isnan(v):
                continue
            if str(v).strip().lower() == 'spending code':
                header_row = i
                break
        else:
            continue
        break
    try:
        if hasattr(uploaded_file, 'seek'):
            uploaded_file.seek(0)
        df = pd.read_excel(uploaded_file, sheet_name=0, header=header_row)
    except Exception as e:
        return None, str(e)
    if df is None or df.empty:
        return None, "The file is empty or could not be read."
    cols = [str(c).strip().lower() for c in df.columns]
    # Find code column
    code_col = None
    for i, c in enumerate(cols):
        if any(k in c for k in ('code', 'var', 'variable', 'category', 'spending')):
            code_col = df.columns[i]
            break
    if code_col is None:
        code_col = df.columns[0]
    # Shared Consumption %
    shared_col = None
    for i, c in enumerate(cols):
        if 'shared' in c and any(k in c for k in ('%', 'pct', 'percent', 'consumption')):
            shared_col = df.columns[i]
            break
    if shared_col is None:
        for i, c in enumerate(cols):
            if 'shared' in c or ('consumption' in c and '%' in c):
                shared_col = df.columns[i]
                break
    # Child Intensity Index
    child_col = None
    for i, c in enumerate(cols):
        if 'child' in c and 'intensity' in c:
            child_col = df.columns[i]
            break
    if child_col is None:
        for i, c in enumerate(cols):
            if 'intensity' in c and 'child' in c:
                child_col = df.columns[i]
                break
    # Vectorized processing: filter valid rows first
    code_series = df[code_col].astype(str).str.strip()
    valid_mask = (code_series.notna()) & (code_series != '') & (code_series != 'nan') & (code_series != 'NAN')
    valid_df = df[valid_mask].copy()
    
    if valid_df.empty:
        return {}, None
    
    # Process codes vectorized
    codes = valid_df[code_col].astype(str).str.strip()
    # Normalize codes: uppercase prefix if it looks like a var code
    codes = codes.apply(lambda x: x[:2].upper() + x[2:] if len(x) >= 2 and x[:2].isalpha() else x)
    
    # Extract values vectorized
    shared_values = valid_df[shared_col] if shared_col is not None else pd.Series([None] * len(valid_df))
    child_values = valid_df[child_col] if child_col is not None else pd.Series([None] * len(valid_df))
    
    # Build output dictionary
    out = {}
    for idx, var_code in codes.items():
        if var_code and var_code != 'NAN':
            out[var_code] = {
                'shared_pct': _to_float(shared_values.loc[idx]) if shared_col is not None else None,
                'child_intensity': _to_float(child_values.loc[idx]) if child_col is not None else None
            }
    return out, None


def load_allocation_from_cache():
    """Load allocation input from the cache file. Returns dict or None."""
    try:
        if ALLOCATION_INPUT_CACHE.exists():
            with open(ALLOCATION_INPUT_CACHE, 'r') as f:
                return json.load(f)
    except Exception:
        pass
    return None


def save_allocation_to_cache(data):
    """Save allocation input to the cache file so it remains valid until replaced."""
    try:
        with open(ALLOCATION_INPUT_CACHE, 'w') as f:
            json.dump(data, f, indent=2)
    except Exception:
        pass


@st.cache_data
def load_default_allocation(alloc_form_filename: str):
    """Load default allocation input from the bundled Excel form."""
    alloc_form_path = Path(__file__).resolve().parent / alloc_form_filename
    if not alloc_form_path.exists():
        return None, f"{alloc_form_filename} not found."
    try:
        parsed, err = parse_allocation_input_excel(BytesIO(alloc_form_path.read_bytes()))
    except Exception as exc:
        return None, str(exc)
    if err:
        return None, err
    return parsed, None


# Load hierarchy structure
@st.cache_data
def load_hierarchy():
    """Load the hierarchy structure from JSON file"""
    try:
        hierarchy_file = Path("hierarchy_structure.json")
        if hierarchy_file.exists():
            with open(hierarchy_file, 'r') as f:
                return json.load(f)
        return None
    except Exception as e:
        st.warning(f"Could not load hierarchy structure: {e}")
        return None

@st.cache_data(ttl=3600)
def load_data(year: int):
    """Load the main dataset. 2019: SAS7BDAT via pyreadstat; 2021: fixed-width TXT."""
    try:
        if year == 2019:
            df, meta = pyreadstat.read_sas7bdat(str(MAIN_FILE))
            df = YEAR_CONFIG.apply_rename(df)
            return df, meta
        else:
            df, meta = load_fixed_width_pumf(MAIN_FILE, LAYOUT_FILE)
            df = YEAR_CONFIG.apply_rename(df)
            return df, meta
    except Exception as e1:
        st.error(f"Error loading main data file: {e1}")
        return None, None

@st.cache_data(ttl=3600)
def load_bootstrap_weights(year: int):
    """Load the bootstrap weights dataset. 2019: SAS7BDAT; 2021: fixed-width TXT."""
    try:
        if year == 2019:
            df_bsw, meta_bsw = pyreadstat.read_sas7bdat(str(BSW_FILE))
            return df_bsw, meta_bsw
        else:
            df_bsw, meta_bsw = load_fixed_width_bsw(BSW_FILE, BSW_LAYOUT_FILE)
            return df_bsw, meta_bsw
    except Exception as e1:
        st.error(f"Error loading bootstrap weights file: {e1}")
        return None, None

def get_bootstrap_weights(df, df_bsw):
    """Get all bootstrap weight column names and merge with main data"""
    if df_bsw is None:
        return df, []
    
    # Merge bootstrap weights with main data on CaseID (handle case differences)
    # Main file uses 'CaseID', bootstrap file uses 'caseid'
    if 'CaseID' in df.columns:
        if 'caseid' in df_bsw.columns:
            df = df.merge(df_bsw, left_on='CaseID', right_on='caseid', how='left')
        elif 'CaseID' in df_bsw.columns:
            df = df.merge(df_bsw, left_on='CaseID', right_on='CaseID', how='left')
    
    # Get all BSW columns
    bsw_cols = [col for col in df.columns if col.startswith('BSW')]
    # Sort numerically by the number after BSW
    def sort_key(col):
        try:
            return int(col.replace('BSW', ''))
        except:
            return 0
    return df, sorted(bsw_cols, key=sort_key)

def get_variable_value(df, var):
    """Get the variable value, handling _C and _D versions.
    If both _C and _D exist, sum them. Otherwise use the available version."""
    var_c = var + '_C'
    var_d = var + '_D'
    
    has_c = var_c in df.columns
    has_d = var_d in df.columns
    
    if has_c and has_d:
        # Both exist, sum them
        return df[var_c].fillna(0) + df[var_d].fillna(0)
    elif has_c:
        # Only _C exists
        return df[var_c]
    elif has_d:
        # Only _D exists
        return df[var_d]
    elif var in df.columns:
        # Base variable exists (no _C or _D)
        return df[var]
    else:
        # Variable doesn't exist
        return pd.Series([np.nan] * len(df), index=df.index)

def calculate_weighted_mean(df, var, weight_col='WeightD'):
    """Calculate weighted mean for a variable, handling _C and _D versions"""
    # Get the variable value (handling _C and _D)
    var_values = get_variable_value(df, var)
    
    # Filter out missing values
    mask = var_values.notna() & (df[weight_col] > 0)
    if mask.sum() == 0:
        return np.nan
    
    weighted_sum = (var_values.loc[mask] * df.loc[mask, weight_col]).sum()
    total_weight = df.loc[mask, weight_col].sum()
    
    if total_weight == 0:
        return np.nan
    
    return weighted_sum / total_weight

def _bootstrap_variance_fast(v, w_main, W_bs, main_estimate):
    """Bootstrap variance using numpy; v, w_main (n,), W_bs (n x B)."""
    if not np.isfinite(main_estimate) or W_bs.size == 0:
        return np.nan
    bs_estimates = []
    for b in range(W_bs.shape[1]):
        m = np.isfinite(v) & (W_bs[:, b] > 0)
        if m.sum() == 0:
            continue
        s = np.sum(v[m] * W_bs[m, b])
        t = np.sum(W_bs[m, b])
        if t > 0:
            e = s / t
            if np.isfinite(e):
                bs_estimates.append(e)
    if len(bs_estimates) == 0:
        return np.nan
    return float(np.mean((np.array(bs_estimates, dtype=np.float64) - main_estimate) ** 2))

def calculate_bootstrap_variance(df, var, weight_col='WeightD', bootstrap_cols=None):
    """Calculate bootstrap variance using bootstrap weights, handling _C and _D versions"""
    if bootstrap_cols is None or len(bootstrap_cols) == 0:
        return np.nan
    
    # Get variable values (handling _C and _D)
    var_values = get_variable_value(df, var)
    
    # Calculate estimate with main weight
    mask = var_values.notna() & (df[weight_col] > 0)
    if mask.sum() == 0:
        return np.nan
    
    weighted_sum = (var_values.loc[mask] * df.loc[mask, weight_col]).sum()
    total_weight = df.loc[mask, weight_col].sum()
    
    if total_weight == 0:
        return np.nan
    
    main_estimate = weighted_sum / total_weight
    
    if np.isnan(main_estimate):
        return np.nan
    
    # Calculate estimates with each bootstrap weight
    bootstrap_estimates = []
    for bs_col in bootstrap_cols:
        if bs_col in df.columns:
            bs_mask = var_values.notna() & (df[bs_col] > 0)
            if bs_mask.sum() > 0:
                bs_weighted_sum = (var_values.loc[bs_mask] * df.loc[bs_mask, bs_col]).sum()
                bs_total_weight = df.loc[bs_mask, bs_col].sum()
                if bs_total_weight > 0:
                    bs_estimate = bs_weighted_sum / bs_total_weight
                    if not np.isnan(bs_estimate):
                        bootstrap_estimates.append(bs_estimate)
    
    if len(bootstrap_estimates) == 0:
        return np.nan
    
    # Bootstrap variance formula: sum((estimate_b - estimate_full)^2) / B
    bootstrap_estimates = np.array(bootstrap_estimates)
    variance = np.mean((bootstrap_estimates - main_estimate) ** 2)
    
    return variance

def calculate_sample_size(df, var, weight_col='WeightD'):
    """Calculate sample size (n) - number of non-missing observations for a variable, handling _C and _D versions"""
    var_values = get_variable_value(df, var)
    
    # Count non-missing observations (where variable has a value and weight > 0)
    mask = var_values.notna() & (df[weight_col] > 0)
    n = mask.sum()
    
    return n

def determine_data_quality_category(cv, n=None, region=None):
    """
    Determine data quality release category (A, E, or F) based on CV, sample size, and region.
    
    According to PUMF User Guide (pages 21-23):
    - Category A (Acceptable): CV <= 16.5%, can be released without restrictions
    - Category E (Marginal): 16.5% < CV <= 33.3%, can be released with warning about high sampling variability
    - Category F (Unacceptable): CV > 33.3%, should not be released
    
    Additional considerations may apply based on sample size and region.
    """
    if np.isnan(cv) or cv is None:
        return 'F'  # Unacceptable if CV cannot be calculated
    
    # Primary classification based on CV thresholds
    if cv <= 16.5:
        quality_category = 'A'
    elif cv <= 33.3:
        quality_category = 'E'
    else:
        quality_category = 'F'
    
    # Additional checks based on sample size (if very small sample, may downgrade)
    # Note: These thresholds may need adjustment based on specific PUMF guide requirements
    if n is not None and n < 10:
        # Very small sample sizes may be downgraded
        if quality_category == 'A':
            quality_category = 'E'  # Downgrade A to E for very small samples
        elif quality_category == 'E':
            quality_category = 'F'  # Downgrade E to F for very small samples
    
    return quality_category

def _canonical_filter_code(x):
    """Normalize categorical codes for comparison (handles 1 vs 01, 2 vs 02 between 2019/2021)."""
    try:
        s = str(x).strip()
        if s.replace(".", "").replace("-", "").isdigit():
            return str(int(float(s)))
        return s
    except (TypeError, ValueError):
        return str(x)


def filter_data(df, filters, income_range=None, income_col='HH_TotInc'):
    """Apply filters to the dataset. Normalizes categorical codes so filters carry across year change (e.g. 2 vs 02)."""
    filtered_df = df.copy()
    
    for var, value in filters.items():
        if value is not None and var in filtered_df.columns:
            if isinstance(value, list):
                if len(value) > 0:
                    # Normalize filter values for 2019/2021 code differences (e.g. "2" vs "02")
                    canon_values = {_canonical_filter_code(v) for v in value}
                    col_canon = filtered_df[var].apply(_canonical_filter_code)
                    filtered_df = filtered_df[col_canon.isin(canon_values)]
            else:
                canon_val = _canonical_filter_code(value)
                col_canon = filtered_df[var].apply(_canonical_filter_code)
                filtered_df = filtered_df[col_canon == canon_val]
    
    # Apply income range filter if provided
    if income_range is not None and income_col in filtered_df.columns:
        min_income, max_income = income_range
        if min_income is not None:
            filtered_df = filtered_df[filtered_df[income_col] >= min_income]
        if max_income is not None:
            filtered_df = filtered_df[filtered_df[income_col] <= max_income]
    
    return filtered_df

def get_unique_values(df, column):
    """Get unique non-null values from a column"""
    if column not in df.columns:
        return []
    unique_vals = df[column].dropna().unique()
    return sorted([v for v in unique_vals if pd.notna(v)])

def format_option_label(var_name, value):
    """Format option label for selectbox"""
    label = format_value(var_name, value)
    return f"{label} ({value})" if label != str(value) else str(value)

def organize_hierarchical_results(results_df, hierarchy_data):
    """Organize results in the same order as the hierarchy Excel file, preserving depth-first nested order."""
    if hierarchy_data is None:
        return None, None
    
    var_to_node = hierarchy_data.get('var_to_node', {})
    level_vars = hierarchy_data.get('level_vars', {})
    hierarchy_order = hierarchy_data.get('hierarchy_order', [])
    
    # Create a mapping of var_code to results (vectorized using to_dict for better performance)
    results_dict = {}
    if not results_df.empty:
        # Convert to dict using records orientation for faster access
        records = results_df.to_dict('records')
        for record in records:
            var_code = record['Spending Code']
            results_dict[var_code] = {
                'var_code': var_code,
                'mean': record['Mean Dollars Per Year'],
                'variance': record['Variance'],
                'std_error': record['Standard Error'],
                'cv': record['Coefficient of Variation'],
                'n': record.get('Sample Size (n)', np.nan),
                'quality': record.get('Data Quality Category', 'F')
            }
    
    hierarchical_results = []
    used = set()
    
    if hierarchy_order:
        # Use exact order from hierarchy (depth-first)
        for var_code in hierarchy_order:
            if var_code in results_dict and var_code not in used:
                used.add(var_code)
                node = var_to_node.get(var_code, {})
                level = node.get('level', 0)
                hierarchical_results.append({
                    'var_code': var_code,
                    'level': level,
                    'parent': node.get('parent'),
                    'description': SPENDING_DESCRIPTIONS.get(var_code, node.get('description', var_code)),
                    'mean': results_dict[var_code]['mean'],
                    'variance': results_dict[var_code]['variance'],
                    'std_error': results_dict[var_code]['std_error'],
                    'cv': results_dict[var_code]['cv'],
                    'n': results_dict[var_code]['n'],
                    'quality': results_dict[var_code]['quality']
                })
        # Append any results not in hierarchy_order (e.g. vars in data but not in Excel)
        for var_code, data in results_dict.items():
            if var_code not in used:
                node = var_to_node.get(var_code, {})
                level = node.get('level', 0)
                hierarchical_results.append({
                    'var_code': var_code,
                    'level': level,
                    'parent': node.get('parent'),
                    'description': SPENDING_DESCRIPTIONS.get(var_code, node.get('description', var_code)),
                    'mean': data['mean'],
                    'variance': data['variance'],
                    'std_error': data['std_error'],
                    'cv': data['cv'],
                    'n': data['n'],
                    'quality': data['quality']
                })
    else:
        # Fallback when hierarchy_order not in JSON: order by level_vars (by level, then file order within level)
        for level in sorted(level_vars.keys(), key=lambda x: int(x) if isinstance(x, str) and x.isdigit() else x):
            for var_code in level_vars[level]:
                if var_code in results_dict:
                    node = var_to_node.get(var_code, {})
                    hierarchical_results.append({
                        'var_code': var_code,
                        'level': int(level) if isinstance(level, str) and str(level).isdigit() else level,
                        'parent': node.get('parent'),
                        'description': SPENDING_DESCRIPTIONS.get(var_code, node.get('description', var_code)),
                        'mean': results_dict[var_code]['mean'],
                        'variance': results_dict[var_code]['variance'],
                        'std_error': results_dict[var_code]['std_error'],
                        'cv': results_dict[var_code]['cv'],
                        'n': results_dict[var_code]['n'],
                        'quality': results_dict[var_code]['quality']
                    })
    
    return hierarchical_results, var_to_node


def _level_from_node(var_to_node, var_code):
    level = var_to_node.get(var_code, {}).get('level')
    if level is None:
        return None
    try:
        return int(level)
    except (TypeError, ValueError):
        return None


def filter_results_by_granularity(hierarchical_results, var_to_node, max_level):
    if max_level is None:
        return hierarchical_results
    # Pre-compute levels for all items to avoid repeated lookups
    levels = [_level_from_node(var_to_node, item['var_code']) for item in hierarchical_results]
    # Filter using list comprehension (faster than appending in loop)
    filtered = [item for item, level in zip(hierarchical_results, levels) if level is None or level <= max_level]
    return filtered


def filter_vars_by_granularity(var_codes, var_to_node, max_level):
    if max_level is None:
        return list(var_codes)
    # Pre-compute levels for all codes and filter using list comprehension
    levels = [_level_from_node(var_to_node, var_code) for var_code in var_codes]
    filtered = [var_code for var_code, level in zip(var_codes, levels) if level is None or level <= max_level]
    return filtered


def _get_direct_children(var_code, hierarchy_order, var_to_node):
    """Return direct children of var_code in the tree, from hierarchy_order (depth-first).
    Direct children are nodes at level L+1 in the segment after var_code until we rise to level <= L."""
    ho = hierarchy_order or []
    idx = next((i for i, c in enumerate(ho) if c == var_code), -1)
    if idx < 0:
        return []
    L = int(var_to_node.get(var_code, {}).get('level') or 0)
    children = []
    i = idx + 1
    while i < len(ho):
        l_raw = var_to_node.get(ho[i], {}).get('level')
        l_i = int(l_raw) if l_raw is not None else None
        if l_i is not None and l_i <= L:
            break
        if l_i is not None and l_i == L + 1:
            children.append(ho[i])
        i += 1
    return children


def compute_granular_allocation(hierarchical_results, var_to_node, hierarchy_data=None):
    """
    Granular Allocation displays a subset of 'Mean Dollars per Year' for items under
    TC001 (Total current consumption) and MG001 (Gifts of money, support payments and
    charitable contributions). For each branch, choose the MOST GRANULAR level such that
    no item at that level has Data Quality 'F': recurse into children only when ALL
    direct children are in results and non-F; if any direct child is missing or has 'F',
    display at this node instead. Values shown are the actual Mean (no scaling).
    """
    if not hierarchical_results:
        return {}
    results_dict = {r['var_code']: r for r in hierarchical_results}
    ho = (hierarchy_data or {}).get('hierarchy_order', [])
    v2n = var_to_node

    def _qual(s):
        return (str(s or '').strip().upper() == 'F')

    def _recurse(node):
        children = _get_direct_children(node, ho, v2n)
        children_in_results = [c for c in children if c in results_dict]
        if not children_in_results:
            # Leaf in our result set (or no children in tree): include node if in results and not F
            if node in results_dict and not _qual(results_dict[node].get('quality')):
                return [node]
            return []
        # Can we go to the deeper level? Only if EVERY direct child (in tree) is in results AND non-F.
        # If any is missing or has F, we must show this node to cover the total.
        any_cannot_go_deeper = any(
            (c not in results_dict) or _qual(results_dict.get(c, {}).get('quality'))
            for c in children
        )
        if any_cannot_go_deeper:
            if node in results_dict and not _qual(results_dict[node].get('quality')):
                return [node]
            return []
        # All children in results and non-F: recurse into each (do not show this node)
        out = []
        for c in children_in_results:
            out.extend(_recurse(c))
        return out

    selected = []
    for root in ('TC001', 'MG001'):
        selected.extend(_recurse(root))

    selected_set = set(selected)
    alloc = {}
    for r in hierarchical_results:
        vc = r['var_code']
        if vc in selected_set:
            m = r.get('mean')
            alloc[vc] = m if (m is not None and not (isinstance(m, float) and np.isnan(m))) else np.nan
        else:
            alloc[vc] = np.nan
    return alloc


def _has_granular_value(ga):
    return ga is not None and not (isinstance(ga, float) and np.isnan(ga))


def _allocation_split(M, shared_pct, child_intensity, n_adults, n_children):
    """Compute Shared Spending, Exclusive Per Child, Exclusive Per Adult so that
    Shared + n_adults*ExclPerAdult + n_children*ExclPerChild = M.
    Child Intensity Index: dollars (of every $10 of exclusive spending) to child; 0=all to adult, 10=all to child.
    shared_pct: fraction 0–1 (if >1 treated as 0–100 and scaled). child_intensity: 0–10.
    """
    s = shared_pct if shared_pct is not None else 0
    if isinstance(s, (int, float)) and s > 1:
        s = s / 100
    s = float(s) if s is not None else 0
    c_idx = child_intensity if child_intensity is not None else 0
    c_idx = max(0, min(10, float(c_idx) if c_idx is not None else 0))
    n_a = max(1, int(n_adults) if n_adults is not None else 1)
    n_c = max(0, int(n_children) if n_children is not None else 0)
    shared = s * M
    excl_total = (1 - s) * M
    if n_c > 0:
        excl_per_child = (c_idx / 10) * excl_total / n_c
        excl_per_adult = ((10 - c_idx) / 10) * excl_total / n_a
    else:
        excl_per_child = 0.0
        excl_per_adult = excl_total / n_a
    return shared, excl_per_child, excl_per_adult


def _force_shared_allocation(allocation_lookup):
    if not allocation_lookup:
        return allocation_lookup
    return {
        var_code: {"shared_pct": 1, "child_intensity": 0}
        for var_code in allocation_lookup.keys()
    }


def _reverse_calculate_allocation(M, shared_d, excl_per_adult_d, excl_per_child_d, n_adults, n_children):
    """Reverse-calculate Shared % and Child Intensity from dollar amounts.
    
    Given total mean M and the calculated shared, exclusive per adult, and exclusive per child amounts,
    determine what Shared % and Child Intensity would produce these values.
    
    Returns: (shared_pct, child_intensity) or (None, None) if calculation is not possible.
    """
    if M is None or M == 0 or np.isnan(M):
        return None, None
    
    # Calculate shared_pct from shared amount
    shared_pct = shared_d / M if shared_d is not None and not np.isnan(shared_d) else None
    
    # Calculate total exclusive spending
    excl_total = 0.0
    if excl_per_adult_d is not None and not np.isnan(excl_per_adult_d) and n_adults > 0:
        excl_total += excl_per_adult_d * n_adults
    if excl_per_child_d is not None and not np.isnan(excl_per_child_d) and n_children > 0:
        excl_total += excl_per_child_d * n_children
    
    # Calculate child_intensity from exclusive amounts
    child_intensity = None
    if excl_total > 0 and n_children > 0 and excl_per_child_d is not None and not np.isnan(excl_per_child_d):
        # excl_per_child = (child_intensity / 10) * excl_total / n_children
        # So: child_intensity = (excl_per_child * n_children / excl_total) * 10
        child_intensity = (excl_per_child_d * n_children / excl_total) * 10
        child_intensity = max(0, min(10, child_intensity))  # Clamp to 0-10
    elif excl_total > 0:
        # If no children, child_intensity is 0
        child_intensity = 0.0
    else:
        # If no exclusive spending, child_intensity is undefined, use 0
        child_intensity = 0.0
    
    # Normalize shared_pct to 0-1 range
    if shared_pct is not None:
        shared_pct = max(0, min(1, shared_pct))
    
    return shared_pct, child_intensity


def compute_aggregated_allocation_factors(hierarchical_results_level7, hierarchical_results_requested, var_to_node, hierarchy_data, n_adults, n_children, original_allocation_lookup):
    """Compute aggregated allocation factors from Level 7 to the requested level.
    
    Args:
        hierarchical_results_level7: Results at Level 7 (most granular)
        hierarchical_results_requested: Results at the requested granularity level
        var_to_node: Variable to node mapping
        hierarchy_data: Hierarchy data structure
        n_adults: Number of adults
        n_children: Number of children
        original_allocation_lookup: Original allocation lookup from Excel (used for Level 7 calculations)
    
    Returns:
        Dictionary mapping var_code to {'shared_pct': float, 'child_intensity': float}
    """
    if not hierarchical_results_level7 or not hierarchical_results_requested:
        return {}
    
    # Create a mapping of var_code to results for quick lookup
    results_level7_dict = {r['var_code']: r for r in hierarchical_results_level7}
    results_requested_dict = {r['var_code']: r for r in hierarchical_results_requested}
    
    # Create aggregated allocation lookup
    aggregated_lookup = {}
    
    # For each node at the requested level, aggregate from its children at Level 7
    for item in hierarchical_results_requested:
        var_code = item['var_code']
        level = int(item.get('level', 0)) if item.get('level') is not None else 0
        
        # If this is already at Level 7, use original allocation values directly
        # (no aggregation needed since Level 7 is the most granular)
        if level >= 6:  # Level 7 is index 6 (0-based) or level 7 (1-based)
            if var_code in original_allocation_lookup:
                aggregated_lookup[var_code] = original_allocation_lookup[var_code].copy()
            # If no original value exists, skip (don't create empty entry)
            continue
        
        # For nodes above Level 7, aggregate from all Level 7 descendants
        shared_pct, child_intensity = _aggregate_allocation_from_children(
            var_code, hierarchical_results_level7, var_to_node, hierarchy_data, 
            n_adults, n_children, original_allocation_lookup
        )
        
        if shared_pct is not None and child_intensity is not None:
            aggregated_lookup[var_code] = {
                'shared_pct': shared_pct,
                'child_intensity': child_intensity
            }
    
    return aggregated_lookup


def _get_all_descendants(var_code, hierarchical_results_level7, var_to_node, hierarchy_data):
    """Get all descendant nodes at Level 7 that are descendants of the given var_code."""
    ho = (hierarchy_data or {}).get('hierarchy_order', [])
    parent_level = int(var_to_node.get(var_code, {}).get('level', 0)) if var_to_node.get(var_code) else 0
    
    descendants = []
    for result in hierarchical_results_level7:
        desc_code = result['var_code']
        desc_level = int(result.get('level', 0)) if result.get('level') is not None else 0
        
        # Check if this node is a descendant by checking if it appears after parent in hierarchy_order
        # and has a higher level number
        if desc_level > parent_level:
            parent_idx = next((i for i, c in enumerate(ho) if c == var_code), -1)
            desc_idx = next((i for i, c in enumerate(ho) if c == desc_code), -1)
            
            if parent_idx >= 0 and desc_idx > parent_idx:
                # Check if this is actually a descendant (not just a later node at same level)
                # by checking if any node between parent and desc has a level <= parent_level
                is_descendant = True
                for i in range(parent_idx + 1, desc_idx):
                    intermediate_code = ho[i]
                    intermediate_level = int(var_to_node.get(intermediate_code, {}).get('level', 0)) if var_to_node.get(intermediate_code) else 0
                    if intermediate_level <= parent_level:
                        is_descendant = False
                        break
                
                if is_descendant:
                    descendants.append(result)
    
    return descendants


def _aggregate_allocation_from_children(parent_var_code, hierarchical_results_level7, var_to_node, hierarchy_data, n_adults, n_children, allocation_lookup):
    """Aggregate allocation values from all Level 7 descendant nodes to a parent node.
    
    Returns: (shared_pct, child_intensity) calculated from weighted average of all Level 7 descendants.
    """
    # Get all Level 7 descendants of this parent node
    descendant_results = _get_all_descendants(parent_var_code, hierarchical_results_level7, var_to_node, hierarchy_data)
    
    if not descendant_results:
        return None, None
    
    # Aggregate Shared $, Exclusive (Adult) $, Exclusive (Child) $ from all descendants
    total_mean = 0.0
    total_shared_d = 0.0
    total_excl_adult_d = 0.0
    total_excl_child_d = 0.0
    
    for descendant in descendant_results:
        desc_code = descendant['var_code']
        desc_mean = descendant.get('mean', 0)
        if desc_mean is None or np.isnan(desc_mean) or desc_mean == 0:
            continue
        
        # Get allocation values for this descendant
        lookup = allocation_lookup.get(desc_code, {}) if allocation_lookup else {}
        shared_pct = lookup.get('shared_pct')
        child_intensity = lookup.get('child_intensity')
        
        # Calculate dollar amounts for this descendant
        shared_d, excl_per_child_d, excl_per_adult_d = _allocation_split(
            desc_mean, shared_pct, child_intensity, n_adults, n_children
        )
        
        # Aggregate using descendant mean as weight
        total_mean += desc_mean
        total_shared_d += shared_d
        total_excl_adult_d += excl_per_adult_d * n_adults
        total_excl_child_d += excl_per_child_d * n_children
    
    if total_mean == 0:
        return None, None
    
    # Calculate per-adult and per-child amounts
    per_adult_d = total_excl_adult_d / n_adults if n_adults > 0 else 0
    per_child_d = total_excl_child_d / n_children if n_children > 0 else 0
    
    # Reverse-calculate Shared % and Child Intensity from aggregated amounts
    shared_pct, child_intensity = _reverse_calculate_allocation(
        total_mean, total_shared_d, per_adult_d, per_child_d, n_adults, n_children
    )
    
    return shared_pct, child_intensity


def build_hierarchical_display(hierarchical_results, var_to_node, hierarchy_data=None, allocation_lookup=None, n_adults=2, n_children=1):
    """Build display data with nested indentation: Level 2 indented from Level 1, Level 3 from Level 2, etc.
    'Granular Allocation': subset of Mean Dollars per Year for TC001/MG001 branches; per branch, the most
    granular level where no item has quality 'F' (if any deeper item has F, show at the parent level).
    'Shared Consumption %' and 'Child Intensity Index': from allocation_lookup, only for rows with Granular Allocation.
    When allocation is loaded: 'Shared Spending', 'Exclusive Spending Per Child', 'Exclusive Spending Per Adult'
    so that Shared + n_adults*ExclPerAdult + n_children*ExclPerChild = Mean Dollars Per Year."""
    display_rows = []
    INDENT_PER_LEVEL = 2  # spaces per hierarchy level
    gran = compute_granular_allocation(hierarchical_results, var_to_node, hierarchy_data)
    alloc = allocation_lookup if allocation_lookup is not None else {}
    
    for item in hierarchical_results:
        q = str(item.get('quality', 'F') or 'F').strip().upper()
        is_f = (q == 'F')
        level = int(item['level']) if item.get('level') is not None else 0
        # Indent each level from the previous: L0=0, L1=2, L2=4, L3=6, ...
        indent = " " * (level * INDENT_PER_LEVEL)
        var_code = item['var_code']
        description = item['description']
        ga = gran.get(var_code, np.nan)
        show_alloc = bool(alloc) and _has_granular_value(ga) and not is_f
        lookup = alloc.get(var_code, {}) if show_alloc else {}
        
        row = {
            'Expenditure Category': f"{indent}{description}",
            'Reported $': "" if is_f else item['mean'],
            'Coefficient of Variation': item.get('cv'),
            'Quality': item.get('quality', 'F'),
            'Allocated $': "" if is_f else ga
        }
        if allocation_lookup is not None:
            if is_f:
                row['Shared %'] = ""
                row['Child Intensity'] = ""
                row['Shared $'] = ""
                row['Exclusive (Adult) $'] = ""
                row['Exclusive (Child) $'] = ""
            else:
                row['Shared %'] = lookup.get('shared_pct') if show_alloc else np.nan
                row['Child Intensity'] = lookup.get('child_intensity') if show_alloc else np.nan
                if show_alloc:
                    shared, excl_per_child, excl_per_adult = _allocation_split(
                        item['mean'],
                        lookup.get('shared_pct'),
                        lookup.get('child_intensity'),
                        n_adults, n_children
                    )
                    row['Shared $'] = shared
                    row['Exclusive (Adult) $'] = excl_per_adult
                    row['Exclusive (Child) $'] = excl_per_child
                else:
                    row['Shared $'] = np.nan
                    row['Exclusive (Adult) $'] = np.nan
                    row['Exclusive (Child) $'] = np.nan
        display_rows.append(row)
    
    return pd.DataFrame(display_rows)

def main():
    banner_path = YEAR_CONFIG.get_banner_path()
    banner_url = os.getenv("SHS_BANNER_URL")
    if banner_url:
        st.image(banner_url, width="stretch")
    elif banner_path.exists():
        st.image(str(banner_path), width="stretch")
    else:
        st.title(f"Survey of Household Spending {year_choice} - Spending Analysis")
    st.markdown(
        """
        <style>
        div[data-testid="stButton"] button[kind="secondary"] {
            background-color: #f0f1f2;
            border-color: #d6d8db;
            color: #1f2933;
            font-size: 0.85rem;
        }
        div[data-testid="stButton"] button[kind="secondary"]:hover {
            background-color: #e6e8ea;
            border-color: #c8ccd1;
            color: #1f2933;
        }
        section.main div[data-testid="stMarkdownContainer"] h3 {
            font-size: 0.55em;
            line-height: 1.2;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )
    
    # Documentation Sidebar
    with st.sidebar:
        st.markdown("### 📚 Documentation")
        doc_tab1, doc_tab2, doc_tab3 = st.tabs(["Quick Start", "Getting Started", "Reference"])
        
        with doc_tab1:
            st.markdown("### Purpose")
            st.markdown("""
            The SHS 2021 Spending Analysis app is a tool for forensic economists used to estimate household expenditures 
            and expenditure allocation factors using Statistics Canada's 2021 Survey of Household Spending (SHS) 
            Public Use Microdata File (PUMF). It is designed to support forensic economic analysis, including 
            fatality and personal injury claims.
            """)
            
            st.markdown("### Access")
            st.markdown("""
            Open the application in a web browser using the current deployment link provided by your team.
            
            No installation is required. Microsoft Excel is only needed if you intend to create or modify custom weight files.
            
            Some analytical functions are restricted to authorised users and require internal authentication when prompted.
            """)
            
            st.markdown("### Basic Workflow")
            st.markdown("""
            **1. Define the Demographic Group**
            
            Use the filters to select the household characteristics that best match the reference household of interest 
            (for example, province, household type, age group, income group).
            
            *Tip: Narrower filters increase specificity but reduce sample size, which can affect data reliability.*
            
            **2. Run the Analysis**
            
            Initiate the calculation to generate:
            - Mean annual household expenditures by category
            - Coefficients of variation (CVs) based on bootstrapping
            
            *Authorised access is required to run calculations.*
            
            **3. Review Data Quality**
            
            Check the CVs alongside the means:
            - Lower CVs indicate more reliable estimates
            - High CVs suggest small samples or unstable results
            
            *If CVs are too high, consider broadening the demographic filters.*
            
            **4. Generate Allocation Factors**
            
            Use the results to obtain expenditure allocation factors (category shares of total household spending). 
            These are suitable for use in court reports and loss calculations.
            """)
            
            st.markdown("### Best Practices")
            st.markdown("""
            - Always assess CVs before relying on results
            - Avoid over-filtering the data
            - Document assumptions clearly when using custom weights
            - Do not include internal access details in reports or shared documents
            """)
        
        with doc_tab2:
            st.markdown("### Introduction")
            st.markdown("""
            The SHS 2021 Spending Analysis web application is an analytical tool for forensic economists. It supports 
            forensic economic analysis using data from Statistics Canada's 2021 Survey of Household Spending (SHS) 
            Public Use Microdata File (PUMF).
            
            The SHS is a nationally representative survey that collects detailed information on household expenditures, 
            demographics, and dwelling characteristics. The PUMF provides anonymised microdata and survey weights that 
            allow analysts to estimate typical household spending patterns for defined demographic groups.
            
            This application is intended to support forensic economists in estimating household expenditures and expenditure 
            allocation factors for use in litigation and advisory work, including fatality and personal injury claims.
            
            The application runs entirely in a web browser. No local installation is required. Microsoft Excel is only 
            needed if users wish to create or modify custom weight files.
            """)
            
            st.markdown("### Access and Authorisation")
            st.markdown("""
            Access the application through the current deployment link provided by your team.
            
            Some functions within the application are restricted to authorised users, including:
            - Running expenditure calculations
            - Generating bootstrap-based coefficients of variation (CVs)
            - Uploading or downloading custom weight or allocation files
            
            When attempting to use these features, users will be prompted to enter an internal access password. 
            This password is provided separately to authorised users and should not be shared or included in external documents.
            """)
            
            st.markdown("### Key Capabilities")
            
            st.markdown("#### 1. Demographic Filtering")
            st.markdown("""
            The application allows users to filter the 2021 SHS data by key household and demographic characteristics, 
            including (but not limited to):
            - Province or region
            - Household type (e.g. single person, couple with children)
            - Age group of the reference person
            - Household income group
            - Housing tenure
            
            These filters define the analytical sample used to calculate expenditure statistics.
            
            **Important:** There is an inherent trade-off between demographic specificity and statistical reliability. 
            Narrow filters reduce sample size, which can increase uncertainty in estimated means. Users should balance 
            precision with data quality.
            """)
            
            st.markdown("#### 2. Reference Household Mapping")
            st.markdown("""
            Filtered demographic groups can be used to approximate a reference household with characteristics similar 
            to those of a household of interest (for example, a deceased individual's household in a fatality claim).
            
            The resulting expenditure profile represents the average spending behaviour of comparable households in the 
            SHS population and may be used as a benchmark in forensic analysis.
            """)
            
            st.markdown("#### 3. Mean Expenditures and Data Quality (CVs)")
            st.markdown("""
            For the selected demographic group, the application calculates:
            - Mean annual household expenditures by category
            - Coefficients of variation (CVs) for each estimate
            
            CVs are derived using a bootstrap methodology consistent with Statistics Canada practice. They provide a 
            direct measure of statistical reliability:
            - Lower CVs indicate more stable estimates
            - Higher CVs indicate greater uncertainty, often due to small sample sizes
            
            Analysts should review CVs carefully and avoid relying on estimates with unacceptably high variability 
            in formal reporting.
            """)
            
            st.markdown("#### 4. Expenditure Categories and Hierarchy")
            st.markdown("""
            Expenditures are organised according to the SHS expenditure hierarchy.
            
            At present, the application's default weighting structure operates at Level 3 of the hierarchy. As a result:
            - Total expenditure allocations are identical whether calculated at Level 3 or aggregated to higher levels
            - Finer reallocations across higher-level groupings require custom weights
            
            This is a temporary limitation and may be expanded in future versions.
            """)
            
            st.markdown("#### 5. Custom Weight Uploads")
            st.markdown("""
            The application supports uploading custom weight files to enable more granular or tailored analyses. 
            This feature is intended for advanced use cases.
            
            Custom weights may be used to:
            - Redistribute expenditures across subcategories
            - Support alternative allocation structures for court reporting
            - Apply bespoke analytical assumptions while preserving total expenditure consistency
            
            Custom weight files are typically prepared in Excel and must follow the required internal format. 
            Uploading or applying custom weights is a password-protected action.
            """)
            
            st.markdown("#### 6. Expenditure Allocation Factors")
            st.markdown("""
            The application can generate expenditure allocation factors, expressed as percentages of total household 
            spending by category.
            
            These allocation factors are commonly used in court reports to:
            - Allocate total household expenditures across categories
            - Support loss calculations and household service valuations
            - Provide transparent, data-driven assumptions
            
            Allocation factors should always be interpreted alongside their associated CVs to ensure sufficient data quality.
            """)
            
            st.markdown("### Best Practice Notes")
            st.markdown("""
            - Always review sample size and CVs before relying on results
            - Broaden demographic filters if estimates appear unstable
            - Use custom weights cautiously and document assumptions clearly
            - Do not include internal access credentials in reports or shared files
            """)
            
            st.markdown("### Conclusion")
            st.markdown("""
            The SHS 2021 Spending Analysis app provides forensic economists with a robust, defensible framework for estimating 
            household expenditures based on nationally representative survey data. By combining demographic filtering, 
            weighted means, and bootstrap-based measures of reliability, the application supports high-quality forensic 
            economic analysis suitable for litigation and advisory work.
            
            For questions about methodology, weighting structures, or appropriate use in reports, consult appropriate 
            guidance or a senior analyst.
            """)
        
        with doc_tab3:
            st.markdown("### Methodology")
            st.markdown("""
            **Bootstrap Variance Estimation**
            
            The bootstrap variance is calculated using the standard Statistics Canada methodology:
            1. Calculate the estimate using the main household weight (WEIGHTD)
            2. Calculate estimates using each of the 1000 bootstrap weights (BSW1 to BSW1000)
            3. Calculate variance as: `Variance = (1/(B-1)) * sum((estimate_b - mean(estimate_b))^2)` where b ranges 
               over all bootstrap replicates
            
            This provides proper variance estimates that account for the complex survey design.
            """)
            
            st.markdown("### Data Quality Flags")
            st.markdown("""
            Each expenditure estimate includes a data quality flag based on the coefficient of variation (CV):
            
            - **A = Publish** (C.V. < 16.6%): Reliable estimate suitable for publication
            - **E = Use with Caution** (16.6% ≤ CV < 35%): Estimate has moderate uncertainty
            - **F = Suppress** (CV ≥ 35%): Estimate has high uncertainty; numeric values are suppressed (shown as blank)
            
            For F-quality items, the row is displayed with the Quality flag and CV shown, but all numeric expenditure 
            values are blank to indicate they should not be used in analysis.
            """)
            
            st.markdown("### Allocation Factors")
            st.markdown("""
            Expenditure allocation factors represent the percentage of total household spending allocated to each 
            expenditure category. These factors are calculated based on:
            - Mean expenditures for the selected demographic group
            - Total household consumption and gifts (TC001 + MG001)
            - Custom allocation inputs (Shared % and Child Intensity) when provided
            
            Allocation factors are used in forensic economic analysis to:
            - Distribute total household spending across categories
            - Support loss calculations
            - Provide defensible assumptions in court reports
            """)
            
            st.markdown("### Custom Weights (Advanced)")
            st.markdown("""
            Custom weight files may be uploaded to support more granular or tailored analyses.
            
            **Default weights operate at Level 3 of the expenditure hierarchy.**
            
            Uploading or applying custom weights requires authorised access.
            
            Custom weights allow analysts to:
            - Redistribute expenditures across subcategories beyond Level 3
            - Apply alternative allocation structures
            - Incorporate bespoke analytical assumptions
            
            Custom weight files must follow the required internal format and are typically prepared in Excel.
            """)
            
            st.markdown("### Current Limitations")
            st.markdown("""
            At present, default weights are available only at Level 3. As a result, total expenditure allocations 
            are the same whether calculated at Level 3 or aggregated to higher levels.
            
            Finer reallocations across higher-level groupings require custom weights.
            """)
            
            st.markdown("### Technical Notes")
            st.markdown("""
            - The application uses all 1000 bootstrap weights (BSW1 to BSW1000) for variance estimation
            - Calculations may take a few minutes when processing all spending categories with bootstrap variance
            - The application uses caching to speed up data loading
            - Ensure sufficient memory for large datasets
            - Spending estimates are in dollars per year (annual household spending)
            - All estimates are weighted using the main household weight (WEIGHTD)
            """)
    
    # Load data and hierarchy
    with st.spinner("Loading data..."):
        df, meta = load_data(year_choice)
        df_bsw, meta_bsw = load_bootstrap_weights(year_choice)
        hierarchy_data = load_hierarchy()
    
    if df is None:
        st.error("Failed to load data. Please check that the data files are in the correct location.")
        return
    
    # Merge bootstrap weights
    df, bootstrap_cols = get_bootstrap_weights(df, df_bsw)
    
    # Filters at the top of the page
    st.header("Select Attributes")
    
    # Get unique values for filter variables
    filters = {}
    income_range = None  # Initialize income range
    
    # Create columns for filters: Left, Middle, Right
    col1, col2, col3 = st.columns(3, gap="large")
    
    # LEFT COLUMN: Geography, Household Characteristics
    _col_prov = YEAR_CONFIG.get_filter_column('PROV')
    _col_hhtype = YEAR_CONFIG.get_filter_column('HHTYPE6')
    _col_hhsize = YEAR_CONFIG.get_filter_column('HHSIZE')
    _col_dwell = YEAR_CONFIG.get_filter_column('DWELTYP')
    _col_tenure = YEAR_CONFIG.get_filter_column('TENURE')
    _col_income = YEAR_CONFIG.get_income_column()
    with col1:
        st.subheader("Geography")
        provinces = get_unique_values(df, _col_prov)
        if provinces:
            selected_provinces = st.multiselect(
                "Province",
                options=provinces,
                format_func=lambda x: format_option_label('PROV', x),
                help="Select one or more provinces. Leave empty to include all."
            )
            if len(selected_provinces) > 0:
                filters[_col_prov] = selected_provinces
        
        st.subheader("Household Characteristics")
        hh_types = get_unique_values(df, _col_hhtype)
        if hh_types:
            selected_hhtype = st.multiselect(
                "Household type",
                options=hh_types,
                format_func=lambda x: format_option_label('HHTYPE6', x),
                help="Select one or more household types. Leave empty to include all."
            )
            if len(selected_hhtype) > 0:
                filters[_col_hhtype] = selected_hhtype
        
        hh_sizes = get_unique_values(df, _col_hhsize)
        if hh_sizes:
            selected_hhsize = st.multiselect(
                "Household size",
                options=hh_sizes,
                format_func=lambda x: format_option_label('HHSIZE', x),
                help="Select one or more household sizes. Leave empty to include all."
            )
            if len(selected_hhsize) > 0:
                filters[_col_hhsize] = selected_hhsize
        
        dwelling_types = get_unique_values(df, _col_dwell)
        if dwelling_types:
            selected_dwell = st.multiselect(
                "Type of dwelling",
                options=dwelling_types,
                format_func=lambda x: format_option_label('DWELTYP', x),
                help="Select one or more dwelling types. Leave empty to include all."
            )
            if len(selected_dwell) > 0:
                filters[_col_dwell] = selected_dwell
        
        tenure = get_unique_values(df, _col_tenure)
        if tenure:
            selected_tenure = st.multiselect(
                "Dwelling tenure",
                options=tenure,
                format_func=lambda x: format_option_label('TENURE', x),
                help="Select one or more tenure types. Leave empty to include all."
            )
            if len(selected_tenure) > 0:
                filters[_col_tenure] = selected_tenure
        
        # Household Total Income Range Slider
        if _col_income in df.columns:
            st.markdown("---")
            st.subheader("Household Total Income Range")
            income_min = float(df[_col_income].min())
            income_max = float(df[_col_income].max())
            # Default to full range (no filtering by default)
            income_default_min = income_min
            income_default_max = income_max
            
            income_options = list(range(int(income_min), int(income_max) + 1, 1000))
            if income_options[-1] != int(income_max):
                income_options.append(int(income_max))
            income_range = st.select_slider(
                "Total Household Income ($)",
                options=income_options,
                value=(int(income_default_min), int(income_default_max)),
                format_func=lambda x: f"${x:,.0f}",
                help="Select the minimum and maximum household income range. Drag the sliders to adjust. Default includes all households."
            )
            st.markdown(f"<p style='font-size: 1em; font-weight: normal;'>Selected range: <strong>${income_range[0]:,.0f}</strong> to <strong>${income_range[1]:,.0f}</strong></p>", unsafe_allow_html=True)
            count_placeholder = st.empty()
            button_cols = st.columns([6, 4])
            with button_cols[0]:
                st.markdown("<div style='margin-top: 1rem;'></div>", unsafe_allow_html=True)
                quintile_cutoffs_btn = st.button(
                    "Show/Hide Quintiles",
                    type="secondary",
                    use_container_width=True,
                    key="quintile_btn_col1"
                )
        else:
            income_range = None
            quintile_cutoffs_btn = False
            count_placeholder = None
            st.session_state.show_quintile_cutoffs = False
    
    _col_rp_age = YEAR_CONFIG.get_filter_column('RP_AGEGRP')
    _col_rp_gender = YEAR_CONFIG.get_filter_column('RP_GENDER')
    _col_rp_marstat = YEAR_CONFIG.get_filter_column('RP_MARSTAT')
    _col_rp_educ = YEAR_CONFIG.get_filter_column('RP_EDUC')
    _col_majinc = YEAR_CONFIG.get_filter_column('HH_MAJINCSRC')
    _col_vehicle = YEAR_CONFIG.get_filter_column('VEHICLEYN')
    _col_sp_age = YEAR_CONFIG.get_filter_column('SP_AGEGRP')
    _col_sp_educ = YEAR_CONFIG.get_filter_column('SP_EDUC')
    _col_p0to4 = YEAR_CONFIG.get_filter_column('P0TO4YN')
    _col_p5to15 = YEAR_CONFIG.get_filter_column('P5TO15YN')
    # MIDDLE COLUMN: Reference Person Demographics
    with col2:
        st.subheader("Reference Person Demographics")
        rp_age = get_unique_values(df, _col_rp_age)
        if rp_age:
            selected_rp_age = st.multiselect(
                "Reference person - Age group",
                options=rp_age,
                format_func=lambda x: format_option_label('RP_AGEGRP', x),
                help="Select one or more age groups. Leave empty to include all."
            )
            if len(selected_rp_age) > 0:
                filters[_col_rp_age] = selected_rp_age
        
        rp_gender = get_unique_values(df, _col_rp_gender)
        if rp_gender:
            selected_rp_gender = st.multiselect(
                "Reference person - Gender",
                options=rp_gender,
                format_func=lambda x: format_option_label('RP_GENDER', x),
                help="Select one or more gender categories. Leave empty to include all."
            )
            if len(selected_rp_gender) > 0:
                filters[_col_rp_gender] = selected_rp_gender
        
        rp_marstat = get_unique_values(df, _col_rp_marstat)
        if rp_marstat:
            selected_rp_marstat = st.multiselect(
                "Reference person - Marital status",
                options=rp_marstat,
                format_func=lambda x: format_option_label('RP_MARSTAT', x),
                help="Select one or more marital statuses. Leave empty to include all."
            )
            if len(selected_rp_marstat) > 0:
                filters[_col_rp_marstat] = selected_rp_marstat
        
        rp_educ = get_unique_values(df, _col_rp_educ)
        if rp_educ:
            selected_rp_educ = st.multiselect(
                "Reference person - Education",
                options=rp_educ,
                format_func=lambda x: format_option_label('RP_EDUC', x),
                help="Select one or more education levels. Leave empty to include all."
            )
            if len(selected_rp_educ) > 0:
                filters[_col_rp_educ] = selected_rp_educ
        
        st.subheader("Major Income Source")
        hh_majinc = get_unique_values(df, _col_majinc)
        if hh_majinc:
            selected_inc = st.multiselect(
                "Household - Major source of income",
                options=hh_majinc,
                format_func=lambda x: format_option_label('HH_MAJINCSRC', x),
                help="Select one or more income sources. Leave empty to include all."
            )
            if len(selected_inc) > 0:
                filters[_col_majinc] = selected_inc
        
        st.subheader("Private Vehicle?")
        vehicle_yn = get_unique_values(df, _col_vehicle)
        if vehicle_yn:
            selected_vehicle = st.multiselect(
                "Owned, leased or operated a vehicle",
                options=vehicle_yn,
                format_func=lambda x: format_option_label('VEHICLEYN', x),
                help="Select one or more options. Leave empty to include all."
            )
            if len(selected_vehicle) > 0:
                filters[_col_vehicle] = selected_vehicle
    
    # RIGHT COLUMN: Spouse, Children, Allocation
    with col3:
        st.subheader("Spouse Demographics")
        # Check if SPOUSEYN exists, otherwise infer from SP_AgeGrp (if it has "96" = No spouse)
        sp_age = get_unique_values(df, _col_sp_age)
        if sp_age:
            selected_sp_age = st.multiselect(
                "Spouse - Age group",
                options=sp_age,
                format_func=lambda x: format_option_label('SP_AGEGRP', x),
                help="Select one or more age groups. Leave empty to include all."
            )
            if len(selected_sp_age) > 0:
                filters[_col_sp_age] = selected_sp_age
        
        sp_educ = get_unique_values(df, _col_sp_educ)
        if sp_educ:
            selected_sp_educ = st.multiselect(
                "Spouse - Education",
                options=sp_educ,
                format_func=lambda x: format_option_label('SP_EDUC', x),
                help="Select one or more education levels. Leave empty to include all."
            )
            if len(selected_sp_educ) > 0:
                filters[_col_sp_educ] = selected_sp_educ
        
        st.subheader("Children in Household")
        p0to4 = get_unique_values(df, _col_p0to4)
        if p0to4:
            selected_p0to4 = st.multiselect(
                "Presence of persons aged 0 to 4 years",
                options=p0to4,
                format_func=lambda x: format_option_label('P0TO4YN', x),
                help="Select one or more options. Leave empty to include all."
            )
            if len(selected_p0to4) > 0:
                filters[_col_p0to4] = selected_p0to4
        
        p5to15 = get_unique_values(df, _col_p5to15)
        if p5to15:
            selected_p5to15 = st.multiselect(
                "Presence of persons aged 5 to 15 years",
                options=p5to15,
                format_func=lambda x: format_option_label('P5TO15YN', x),
                help="Select one or more options. Leave empty to include all."
            )
            if len(selected_p5to15) > 0:
                filters[_col_p5to15] = selected_p5to15
        
        st.markdown(
            """
            <style>
                .allocation-highlight {
                    background-color: var(--warning-background-color, #fff3cd);
                    padding: 0.35rem 0.5rem;
                    border-radius: 0.35rem;
                    display: inline-block;
                }
            </style>
            """,
            unsafe_allow_html=True,
        )
        st.markdown(
            """
            <div class="allocation-highlight"><strong>Allocation: Household Composition *</strong></div>
            """,
            unsafe_allow_html=True,
        )
        total_adults = st.selectbox(
            "Total Adults *",
            options=["— Select —", 1, 2, 3, 4],
            index=0,
            key="total_adults",
            help="Required. Number of adults (1–4). Select a value to enable Calculate."
        )
        total_children = st.selectbox(
            "Total Children *",
            options=["— Select —", 0, 1, 2, 3, 4, 5, 6],
            index=0,
            key="total_children",
            help="Required. Number of children (0–6). Select a value to enable Calculate."
        )
        st.markdown(
            """
            <div class="allocation-highlight"><em>Required before Calculate. You must choose a value (do not leave as "— Select —").</em></div>
            """,
            unsafe_allow_html=True,
        )
    
    # Update session state with current filters for real-time updates
    st.session_state.filters = filters
    st.session_state.income_range = income_range
    
    # Invalidate cached results when data year or filters change (force recalculation)
    def _calc_context_snapshot():
        f_items = tuple(sorted((k, tuple(v) if isinstance(v, (list, tuple)) else v) for k, v in filters.items()))
        inc = income_range if income_range is None else (float(income_range[0]), float(income_range[1]))
        return (year_choice, f_items, inc)
    _ctx = _calc_context_snapshot()
    if st.session_state.get("_last_calc_context") != _ctx:
        for key in ("calculation_mode", "results", "hierarchy_data"):
            st.session_state.pop(key, None)
        st.session_state["_last_calc_context"] = _ctx
    
    # Calculate and display matching records count in real-time
    filtered_df = filter_data(df, filters, income_range=st.session_state.income_range, income_col=YEAR_CONFIG.get_income_column())
    filtered_count = len(filtered_df)
    
    # Display matching records count box
    target_placeholder = count_placeholder or st.empty()
    if filtered_count == 0:
        target_placeholder.markdown(
            """
            <div style="font-size:1.2rem; font-weight:600; padding:0.6rem 0.8rem; border-radius:0.5rem; background:#fff3cd; border:1px solid #ffecb5; color:#664d03;">
                0 records match your selected criteria. Please adjust your selections.
            </div>
            """,
            unsafe_allow_html=True,
        )
    else:
        target_placeholder.markdown(
            f"""
            <div style="font-size:1.2rem; font-weight:600; padding:0.6rem 0.8rem; border-radius:0.5rem; background:#e8f4ff; border:1px solid #b6d4fe; color:#084298;">
                <span style="font-size:1.4rem; font-weight:700;">{filtered_count:,}</span> records match your selected criteria.
            </div>
            """,
            unsafe_allow_html=True,
        )
    
    # Store filtered count in session state
    st.session_state.filtered_count = filtered_count
    
    if quintile_cutoffs_btn:
        show_quintile_cutoffs = not st.session_state.get("show_quintile_cutoffs", False)
        st.session_state.show_quintile_cutoffs = show_quintile_cutoffs

    # Quintile cutoffs: run when toggle is on
    if st.session_state.get("show_quintile_cutoffs", False):
        fq = filter_data(df, st.session_state.filters, income_range=st.session_state.get('income_range'), income_col=YEAR_CONFIG.get_income_column())
        if YEAR_CONFIG.get_income_column() not in fq.columns:
            st.error(f"Household total income ({YEAR_CONFIG.get_income_column()}) not found.")
        else:
            inc = fq[YEAR_CONFIG.get_income_column()].dropna()
            w = fq.loc[inc.index, 'WeightD'] if 'WeightD' in fq.columns else pd.Series(1.0, index=inc.index)
            w = w.fillna(0)
            valid = (inc.notna()) & (w > 0)
            inc, w = inc[valid], w[valid]
            if len(inc) == 0:
                st.error("No valid income data in the filtered sample.")
            else:
                ord = np.argsort(inc.values)
                inc_s, w_s = inc.values[ord], w.values[ord]
                cw = np.cumsum(w_s)
                tw = cw[-1]
                cutoffs = []
                for p in [20, 40, 60, 80]:
                    t = tw * (p / 100)
                    i = np.searchsorted(cw, t, side='left')
                    i = min(i, len(inc_s) - 1)
                    cutoffs.append(float(inc_s[i]))
                st.session_state.quintile_cutoffs = cutoffs
                tab = [["Q1–Q2", f"${cutoffs[0]:,.0f}"], ["Q2–Q3", f"${cutoffs[1]:,.0f}"], ["Q3–Q4", f"${cutoffs[2]:,.0f}"], ["Q4–Q5", f"${cutoffs[3]:,.0f}"]]
                table_cols = st.columns([1, 4])
                with table_cols[0]:
                    st.dataframe(pd.DataFrame(tab, columns=["Boundary", "Income"]), use_container_width=True, hide_index=True)
    
    # Main content area
    st.header("Spending Analysis")
    
    if len(filtered_df) == 0:
        return
    
    # Allocation Input: load from cache if not yet in session (remains valid until replaced by new upload)
    if 'allocation_input' not in st.session_state:
        cached = load_allocation_from_cache()
        if isinstance(cached, dict):
            st.session_state['allocation_input'] = cached
        else:
            defaults, default_err = load_default_allocation(YEAR_CONFIG.get_allocation_form_filename())
            st.session_state['allocation_input'] = defaults if isinstance(defaults, dict) else None
            if default_err:
                st.session_state['allocation_input_default_error'] = default_err
    
    st.markdown("**Allocation Input (optional)**")
    default_alloc_error = st.session_state.get("allocation_input_default_error")
    if default_alloc_error:
        st.warning(f"Default allocations could not be loaded: {default_alloc_error}")
    alloc_mode = st.radio(
        "Allocation mode",
        ["Default Allocations", "Custom Allocations"],
        index=0,
        horizontal=True,
        key="allocation_mode",
    )
    if alloc_mode == "Custom Allocations":
        st.caption("Your custom Allocation Input Form will remain valid until browser is closed.")
        alloc_form_path = Path(__file__).resolve().parent / YEAR_CONFIG.get_allocation_form_filename()
        pw_col, _pw_rest = st.columns([1, 4])
        with pw_col:
            pw = st.text_input("Password for Allocation Input Form (download/upload)", type="password", key="alloc_pw")
        if (pw or "") == "CPC123":
            st.session_state["password_verified"] = True
            if alloc_form_path.exists():
                form_bytes = alloc_form_path.read_bytes()
                st.download_button("Download Allocation Input Form (.xlsx)", data=form_bytes, file_name=YEAR_CONFIG.get_allocation_form_filename(), mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="alloc_download")
            else:
                st.caption(f"{YEAR_CONFIG.get_allocation_form_filename()} not found in project root; add it to enable download.")
            uploaded = st.file_uploader("Upload Allocation Input Form (Excel)", type=['xlsx', 'xls'], key="allocation_upload")
            if uploaded is not None:
                parsed, err = parse_allocation_input_excel(uploaded)
                if err:
                    st.error(f"Could not parse Allocation Input file: {err}")
                elif parsed:
                    st.session_state['allocation_input'] = parsed
                    try:
                        save_allocation_to_cache(parsed)
                    except Exception:
                        pass
                    st.success(f"Allocation input loaded for {len(parsed)} categories. It will remain valid until replaced.")

    st.toggle(
        "Hide allocation factors",
        value=False,
        key="hide_allocation_factors",
        help="Turn on to hide allocation factor columns in the expenditure table."
    )
    granularity_cols = st.columns([1, 4])
    with granularity_cols[0]:
        granularity_level = st.slider(
            "Level of Granularity",
            min_value=3,
            max_value=7,
            value=7,
            step=1,
            key="granularity_level",
            help="Select the maximum hierarchy level to include. Level 7 includes the most detailed categories.",
        )
        st.markdown(
            f"<div style='text-align:center; font-size:1.4rem; font-weight:700;'>Level {granularity_level}</div>",
            unsafe_allow_html=True,
        )
    
    # Toggle for calculating allocation factors using lower-level weights
    _hide_allocation_factors = st.session_state.get("hide_allocation_factors", False)
    use_lower_level_weights = st.toggle(
        "Calculate allocation factors using lower-level weights?",
        value=False,
        key="use_lower_level_weights",
        disabled=_hide_allocation_factors,
        help="When enabled, calculates allocation factors by aggregating from the most granular level (Level 7) up to the selected granularity level, using expenditure amounts as weights."
    )
    if _hide_allocation_factors and st.session_state.get("use_lower_level_weights", False):
        st.session_state["use_lower_level_weights"] = False
        st.info("This option is only available when 'Hide allocation factors' is turned off.")

    st.markdown("---")
    
    # Calculate: show when adults/children set or allocation factors are hidden
    _ta = st.session_state.get("total_adults", "— Select —")
    _tc = st.session_state.get("total_children", "— Select —")
    _adults_ok = _ta != "— Select —" and _ta in [1, 2, 3, 4]
    _children_ok = _tc != "— Select —" and _tc in [0, 1, 2, 3, 4, 5, 6]
    _hide_allocation_factors = st.session_state.get("hide_allocation_factors", False)
    _show_calculate = _hide_allocation_factors or (_adults_ok and _children_ok)
    if _show_calculate:
        btn_col, _ = st.columns([1, 4])
        with btn_col:
            calculate_income_range = st.button("Calculate", type="primary", use_container_width=True)
    else:
        calculate_income_range = False
        btn_col, _ = st.columns([1, 4])
        with btn_col:
            st.warning(
                "Note: To proceed, you must first hide the Allocation Factors, or set the total number "
                "of adults and total number of children."
            )
    
    st.markdown("---")
    
    def _run_calculation():
        _ta = st.session_state.get("total_adults", "— Select —")
        _tc = st.session_state.get("total_children", "— Select —")
        _hide_allocation_factors = st.session_state.get("hide_allocation_factors", False)
        st.session_state["force_shared_allocation"] = bool(_hide_allocation_factors)
        if not _hide_allocation_factors:
            if _ta == "— Select —" or _ta not in [1, 2, 3, 4]:
                st.error("Please select **Total Adults** (1–4) in Allocation: Household Composition before calculating.")
                return False
            if _tc == "— Select —" or _tc not in [0, 1, 2, 3, 4, 5, 6]:
                st.error("Please select **Total Children** (0–6) in Allocation: Household Composition before calculating.")
                return False
        if _ta != "— Select —" and _ta in [1, 2, 3, 4]:
            st.session_state["allocation_n_adults"] = int(_ta)
        if _tc != "— Select —" and _tc in [0, 1, 2, 3, 4, 5, 6]:
            st.session_state["allocation_n_children"] = int(_tc)
        if len(bootstrap_cols) == 0:
            st.error("No bootstrap weights found in the dataset. Cannot calculate variance estimates.")
            return False
        filtered_df = filter_data(df, st.session_state.filters, income_range=st.session_state.income_range, income_col=YEAR_CONFIG.get_income_column())
        st.info(f"Using {len(bootstrap_cols)} bootstrap weights for variance estimation.")
        overall_progress_bar = st.progress(0)
        overall_status_text = st.empty()
        overall_status_text.text("Phase 1 of 2: Calculating individual spending estimates...")
        results = []
        def variable_exists(df, var):
            return (var in df.columns or (var + '_C') in df.columns or (var + '_D') in df.columns)
        ho = hierarchy_data.get('hierarchy_order', []) if hierarchy_data else []
        var_to_node = hierarchy_data.get('var_to_node', {}) if hierarchy_data else {}
        granularity_level = int(st.session_state.get("granularity_level", 7))
        max_granularity_level = granularity_level - 1
        all_hierarchy_vars = ho if ho else sorted(set(ALL_SPENDING_VARS) | PARENT_TOTALS)
        available_spending_vars = [var for var in all_hierarchy_vars if variable_exists(filtered_df, var)]
        available_spending_vars = filter_vars_by_granularity(
            available_spending_vars,
            var_to_node,
            max_granularity_level
        )
        if len(available_spending_vars) == 0:
            st.error("No spending variables found in the dataset.")
            overall_progress_bar.empty()
            overall_status_text.empty()
            return False
        var_to_category = {}
        for cat, vars_list in SPENDING_CATEGORIES.items():
            for var in vars_list:
                var_to_category[var] = cat
        
        # Pre-compute variable column mappings to avoid repeated column checks
        var_column_map = {}
        for var in available_spending_vars:
            var_c = var + '_C'
            var_d = var + '_D'
            has_c = var_c in filtered_df.columns
            has_d = var_d in filtered_df.columns
            has_base = var in filtered_df.columns
            var_column_map[var] = {'has_c': has_c, 'has_d': has_d, 'has_base': has_base}
        
        w_main = np.asarray(filtered_df['WeightD'], dtype=np.float64)
        W_bootstrap = filtered_df[bootstrap_cols].to_numpy(dtype=np.float64)
        progress_bar = st.progress(0)
        status_text = st.empty()
        total_vars = len(available_spending_vars)
        for idx, var in enumerate(available_spending_vars):
            status_text.text(f"Processing {var} ({idx + 1}/{total_vars})...")
            # Use pre-computed column mapping for faster access
            col_info = var_column_map[var]
            if col_info['has_c'] and col_info['has_d']:
                var_vals = filtered_df[var + '_C'].fillna(0) + filtered_df[var + '_D'].fillna(0)
            elif col_info['has_c']:
                var_vals = filtered_df[var + '_C']
            elif col_info['has_d']:
                var_vals = filtered_df[var + '_D']
            elif col_info['has_base']:
                var_vals = filtered_df[var]
            else:
                var_vals = pd.Series([np.nan] * len(filtered_df), index=filtered_df.index)
            v = np.asarray(var_vals, dtype=np.float64)
            mask_main = np.isfinite(v) & (w_main > 0)
            n = int(np.sum(mask_main))
            if n == 0:
                mean_est = np.nan
                variance = np.nan
            else:
                mean_est = float(np.sum(v[mask_main] * w_main[mask_main]) / np.sum(w_main[mask_main]))
                variance = _bootstrap_variance_fast(v, w_main, W_bootstrap, mean_est)
            std_error = np.sqrt(variance) if np.isfinite(variance) else np.nan
            cv = (std_error / mean_est * 100) if np.isfinite(mean_est) and mean_est != 0 else np.nan
            _col_prov = YEAR_CONFIG.get_filter_column('PROV')
            region = filtered_df[_col_prov].mode().iloc[0] if _col_prov in filtered_df.columns and len(filtered_df[_col_prov].mode()) > 0 else None
            quality_category = determine_data_quality_category(cv, n=n, region=region)
            category = var_to_category.get(var, "Other")
            spending_desc = SPENDING_DESCRIPTIONS.get(var, "Spending description not available")
            results.append({
                'Spending Code': var, 'Spending Description': spending_desc, 'Spending Category': category,
                'Mean Dollars Per Year': mean_est, 'Variance': variance, 'Standard Error': std_error,
                'Coefficient of Variation': cv, 'Sample Size (n)': n, 'Data Quality Category': quality_category
            })
            progress_bar.progress((idx + 1) / total_vars)
            overall_progress_bar.progress(0.7 * (idx + 1) / total_vars)
        st.session_state.results = pd.DataFrame(results)
        st.session_state.hierarchy_data = hierarchy_data
        progress_bar.empty()
        status_text.empty()
        overall_progress_bar.progress(1.0)
        avg_household_income = calculate_weighted_mean(filtered_df, YEAR_CONFIG.get_income_column())
        avg_income_variance = calculate_bootstrap_variance(filtered_df, YEAR_CONFIG.get_income_column(), bootstrap_cols=bootstrap_cols)
        avg_income_se = np.sqrt(avg_income_variance) if not np.isnan(avg_income_variance) else np.nan
        tc001_values = get_variable_value(filtered_df, 'TC001')
        if tc001_values.notna().any():
            avg_current_consumption = calculate_weighted_mean(filtered_df, 'TC001')
            avg_consumption_variance = calculate_bootstrap_variance(filtered_df, 'TC001', bootstrap_cols=bootstrap_cols)
            avg_consumption_se = np.sqrt(avg_consumption_variance) if not np.isnan(avg_consumption_variance) else np.nan
        else:
            avg_current_consumption = np.nan
            avg_consumption_se = np.nan
        st.session_state.avg_household_income = avg_household_income
        st.session_state.avg_income_se = avg_income_se
        st.session_state.avg_current_consumption = avg_current_consumption
        st.session_state.avg_consumption_se = avg_consumption_se
        overall_progress_bar.progress(1.0)
        overall_progress_bar.empty()
        overall_status_text.empty()
        st.success("Calculations complete!")
        st.session_state['calculation_mode'] = "income_range"
        return True
    
    # Calculate: if password already verified (e.g. from alloc form), run directly; else require password
    if calculate_income_range:
        if st.session_state.get("password_verified"):
            if not _run_calculation():
                return
        else:
            st.session_state["pending_calculate"] = True
    
    if st.session_state.get("pending_calculate"):
        pwd_col, _pwd_rest = st.columns([1, 4])
        with pwd_col:
            pwd_calc = st.text_input("Enter password to run calculation", type="password", key="pwd_calc")
        confirm_calc = st.button("Confirm and run", key="confirm_calc")
        if confirm_calc:
            if (pwd_calc or "") != "CPC123":
                st.error("Incorrect password.")
            else:
                st.session_state["pending_calculate"] = False
                st.session_state["password_verified"] = True
                if not _run_calculation():
                    return
    
    # Display results based on calculation mode
    calculation_mode_display = st.session_state.get('calculation_mode', None)
    
    # Show regular results only if calculation mode is income_range
    if calculation_mode_display == "income_range" and 'results' in st.session_state and st.session_state.results is not None:
        results_df = st.session_state.results.copy()
        
        # Round numeric columns
        numeric_cols = ['Mean Dollars Per Year', 'Variance', 'Standard Error', 'Coefficient of Variation']
        for col in numeric_cols:
            if col in results_df.columns:
                results_df[col] = results_df[col].round(2)
        
        # Add CSS for subtle table shading and column alignment
        st.markdown("""
        <style>
        /* Subtle shading for dataframes */
        div[data-testid="stDataFrame"] > div {
            background-color: #f8f9fa !important;
        }
        div[data-testid="stDataFrame"] table {
            background-color: #fafbfc !important;
        }
        div[data-testid="stDataFrame"] thead tr th {
            background-color: #f0f1f2 !important;
        }
        /* Left margin metrics and summary table: larger fonts to match pie chart */
        .metrics-margin { font-size: 1rem; background: #f8f9fa; padding: 10px 12px; border-radius: 8px; border: 1px solid #e9ecef; }
        .metrics-margin .metric-label { display: block; color: #6c757d; font-size: 0.95rem; margin-bottom: 2px; }
        .metrics-margin .metric-value { font-size: 1.15rem; font-weight: 600; color: #1e293b; }
        .metrics-margin .metric-item { padding: 8px 0; border-bottom: 1px solid #e9ecef; }
        .metrics-margin .metric-item:last-child { border-bottom: none; padding-bottom: 0; }
        /* Summary allocation table: 2 cols, larger font to match pie */
        .summary-allocation-table { font-size: 1.1rem; border-collapse: collapse; width: 100%; border-radius: 8px; overflow: hidden; box-shadow: 0 1px 3px rgba(0,0,0,0.08); }
        .summary-allocation-table th { background: #5a8fc4; color: #fff; padding: 10px 14px; text-align: left; font-weight: 600; font-size: 1.05rem; }
        .summary-allocation-table td { padding: 10px 14px; border-bottom: 1px solid #e9ecef; background: #fafbfc; font-size: 1.05rem; }
        .summary-allocation-table tr:last-child td { border-bottom: none; }
        .summary-allocation-table td:last-child { text-align: right; font-weight: 500; }
        /* Spending Percentages: prominent block above the table */
        .spending-pct-prominent { font-size: 1.15rem; font-weight: 600; color: #1e293b; background: #e8f4f8; padding: 12px 16px; border-radius: 8px; border: 1px solid #2c3e50; margin-bottom: 12px; }
        .spending-pct-prominent .pct-val { font-size: 1.3rem; font-weight: 700; color: #1a365d; }
        </style>
        <script>
        // Right-justify specific column headers and cells
        function alignNumericColumns() {
            const tables = document.querySelectorAll('div[data-testid="stDataFrame"] table');
            const numericHeaders = ['Reported $', 'Coefficient of Variation', 'Quality', 'Allocated $', 'Shared %', 'Child Intensity', 'Shared $', 'Exclusive (Adult) $', 'Exclusive (Child) $'];
            
            tables.forEach(table => {
                const headers = Array.from(table.querySelectorAll('thead th'));
                headers.forEach((th, colIndex) => {
                    const headerText = th.textContent.trim();
                    if (numericHeaders.includes(headerText)) {
                        th.style.textAlign = 'right';
                        // Also align all cells in this column
                        const rows = table.querySelectorAll('tbody tr');
                        rows.forEach(row => {
                            const cell = row.querySelectorAll('td')[colIndex];
                            if (cell) {
                                cell.style.textAlign = 'right';
                            }
                        });
                    }
                });
            });
        }
        
        // Run immediately and after a delay to catch dynamically loaded tables
        alignNumericColumns();
        setTimeout(alignNumericColumns, 200);
        setTimeout(alignNumericColumns, 500);
        </script>
        """, unsafe_allow_html=True)
        
        # Organize results hierarchically (for summary and table)
        hierarchy_data_display = st.session_state.get('hierarchy_data', hierarchy_data)
        hierarchical_results_full, var_to_node = organize_hierarchical_results(results_df, hierarchy_data_display)
        granularity_level = int(st.session_state.get("granularity_level", 7))
        max_granularity_level = granularity_level - 1
        
        # Get results at Level 7 (most granular) for lower-level weight calculation
        hierarchical_results_level7 = filter_results_by_granularity(
            hierarchical_results_full,
            var_to_node,
            6  # Level 7 is max level (0-based: 6)
        ) if hierarchical_results_full else []
        
        # Get results at requested granularity level
        hierarchical_results = filter_results_by_granularity(
            hierarchical_results_full,
            var_to_node,
            max_granularity_level
        )
        
        gran_alloc = compute_granular_allocation(hierarchical_results, var_to_node, hierarchy_data_display) if hierarchical_results else {}
        allocation_display = st.session_state.get('allocation_input')
        force_shared_allocation = st.session_state.get("force_shared_allocation", False)
        hide_allocation_factors = st.session_state.get("hide_allocation_factors", False)
        use_lower_level_weights = st.session_state.get("use_lower_level_weights", False) and not hide_allocation_factors
        
        # If using lower-level weights, compute aggregated allocation factors from Level 7
        # IMPORTANT: When switch is OFF, use ONLY original spreadsheet values (no aggregation)
        # When switch is ON, use ONLY aggregated values (completely replace spreadsheet values)
        if use_lower_level_weights and hierarchical_results_level7 and hierarchical_results:
            # Get original allocation for Level 7 calculations
            original_allocation_calc = _force_shared_allocation(allocation_display) if force_shared_allocation else allocation_display
            
            # Compute aggregated allocation factors from Level 7 up to requested level
            aggregated_allocation = compute_aggregated_allocation_factors(
                hierarchical_results_level7,
                hierarchical_results,
                var_to_node,
                hierarchy_data_display,
                int(st.session_state.get('allocation_n_adults', 2)),
                int(st.session_state.get('allocation_n_children', 0)),
                original_allocation_calc if original_allocation_calc else {}
            )
            
            # When switch is ON, use ONLY aggregated values (completely replace, don't merge)
            # Only include nodes that have aggregated values
            allocation_calc = {}
            for item in hierarchical_results:
                var_code = item['var_code'] if isinstance(item, dict) else item
                if var_code in aggregated_allocation:
                    allocation_calc[var_code] = aggregated_allocation[var_code].copy()
        else:
            # When switch is OFF, use ONLY original spreadsheet values
            allocation_calc = _force_shared_allocation(allocation_display) if force_shared_allocation else allocation_display
        
        n_a = int(st.session_state.get('allocation_n_adults', 2))
        n_c = int(st.session_state.get('allocation_n_children', 0))
        hide_allocation_factors = st.session_state.get("hide_allocation_factors", False)
        
        # Summary block (same as Excel): Total Consumption and Gifts, N Adults/Children, Shared/Exclusive with Dollars|Percent
        total_consumption_gifts = 0
        if 'Spending Code' in results_df.columns and 'Mean Dollars Per Year' in results_df.columns:
            total_consumption_gifts = (results_df.loc[results_df['Spending Code'] == 'TC001', 'Mean Dollars Per Year'].sum() +
                results_df.loc[results_df['Spending Code'] == 'MG001', 'Mean Dollars Per Year'].sum())
        total_shared_d, total_excl_adult_d, total_excl_child_d = 0.0, 0.0, 0.0
        if allocation_calc and hierarchical_results:
            for item in hierarchical_results:
                ga = gran_alloc.get(item['var_code'], np.nan)
                if not _has_granular_value(ga):
                    continue
                lookup = allocation_calc.get(item['var_code'], {})
                v1, v2 = lookup.get('shared_pct'), lookup.get('child_intensity')
                shared, excl_c, excl_a = _allocation_split(item['mean'], v1, v2, n_a, n_c)
                total_shared_d += shared
                total_excl_adult_d += n_a * excl_a
                total_excl_child_d += n_c * excl_c
        total_alloc = total_shared_d + total_excl_adult_d + total_excl_child_d
        pct_shared = (total_shared_d / total_alloc * 100) if total_alloc else 0
        per_adult_d = (total_excl_adult_d / n_a) if n_a else 0
        per_child_d = (total_excl_child_d / n_c) if n_c else 0
        pct_per_adult = (per_adult_d / total_alloc * 100) if total_alloc else 0
        pct_per_child = (per_child_d / total_alloc * 100) if total_alloc else 0
        pct_agg_adults = (total_excl_adult_d / total_alloc * 100) if total_alloc else 0
        pct_agg_children = (total_excl_child_d / total_alloc * 100) if total_alloc else 0
        
        # Web-only layout: stacked metrics in left margin (smaller font), then table (left) + pie (right) with 2-col table
        margin_col, main_col = st.columns([1, 4])
        with margin_col:
            _v1 = f"${round(total_consumption_gifts, 0):,.0f}" if total_consumption_gifts else "—"
            _v2 = str(int(n_a))
            _v3 = str(int(n_c))
            st.markdown(f'''
            <div class="metrics-margin">
              <div class="metric-item"><span class="metric-label">Total Consumption and Gifts</span><span class="metric-value">{_v1}</span></div>
              <div class="metric-item"><span class="metric-label">Number of Adults</span><span class="metric-value">{_v2}</span></div>
              <div class="metric-item"><span class="metric-label">Number of Children</span><span class="metric-value">{_v3}</span></div>
            </div>''', unsafe_allow_html=True)
        with main_col:
            if not hide_allocation_factors:
                # Table: 2 columns only — Labels (with % in label) and Dollars
                if allocation_calc and total_alloc:
                    lbl1 = f"Shared Spending = {pct_shared:.2f}%"
                    lbl2 = f"Exclusive Spending: Adult(s) = {n_a} × {pct_per_adult:.2f}% = {pct_agg_adults:.2f}%"
                    lbl3 = f"Exclusive Spending: Child(ren) = {n_c} × {pct_per_child:.2f}% = {pct_agg_children:.2f}%"
                    d1 = f"${round(total_shared_d, 0):,.0f}"
                    d2 = f"${round(per_adult_d, 0):,.0f}"
                    d3 = f"${round(per_child_d, 0):,.0f}"
                else:
                    lbl1, lbl2, lbl3 = "Shared Spending", "Exclusive Spending per Adult", "Exclusive Spending per Child"
                    d1 = d2 = d3 = "—"
                summary_table_html = f'''
                <table class="summary-allocation-table">
                  <thead><tr><th>Allocation of Total Consumption and Gifts</th><th>Dollars</th></tr></thead>
                  <tbody>
                    <tr><td>{lbl1}</td><td>{d1}</td></tr>
                    <tr><td>{lbl2}</td><td>{d2}</td></tr>
                    <tr><td>{lbl3}</td><td>{d3}</td></tr>
                  </tbody>
                </table>'''
                tab_col, pie_col = st.columns(2)
                with tab_col:
                    if allocation_calc and total_alloc:
                        pct_block = f'''<div class="spending-pct-prominent">Shared <span class="pct-val">{pct_shared:.2f}%</span> · Exclusive – Adults: <span class="pct-val">{pct_agg_adults:.2f}%</span> · Exclusive – Children: <span class="pct-val">{pct_agg_children:.2f}%</span></div>'''
                        st.markdown(pct_block, unsafe_allow_html=True)
                    st.markdown(summary_table_html, unsafe_allow_html=True)
                with pie_col:
                    if allocation_calc and total_alloc and total_alloc > 0:
                        pie_labels = ["Shared"] + [f"Adult {i+1}" for i in range(n_a)] + [f"Child {i+1}" for i in range(n_c)]
                        pie_values = [total_shared_d] + [per_adult_d] * n_a + [per_child_d] * n_c
                        try:
                            import plotly.graph_objects as go
                            fig = go.Figure(data=[go.Pie(
                                labels=pie_labels,
                                values=pie_values,
                                hole=0.35,
                                textinfo="none",
                                textposition="outside",
                                texttemplate="%{label}<br>%{percent:.2%}",
                                textfont=dict(size=13, color="#1f2933")
                            )])
                            fig.update_layout(margin=dict(t=24, b=60, l=20, r=20), height=240, showlegend=False, font=dict(size=13))
                            st.plotly_chart(fig, use_container_width=True)
                        except Exception:
                            pass
        
        # Display by expenditure category (same columns as Excel)
        st.subheader(f"Allocation by Expenditure Category ({year_choice})")
        _quality_help = "**Quality:** A = Publish (C.V.<16.6%); E = Use with Caution (16.6%≤CV<35%); F = Suppress (CV≥35%)."
        _child_help = "**Child Intensity:** For every ten units of exclusive spending in a home containing adults and children, a 0.00 score means that 100% is spent on a representative adult, a 1.00 score means that 100% is spent on a representative child, and a 0.50 score means equal spending on children and adults."
        def _fmt(x, fmt):
            if pd.isna(x) or (isinstance(x, (int, float)) and (x != x)): return ""
            if fmt == 'cur': return f"${float(x):,.2f}" if (isinstance(x, (int, float)) or np.issubdtype(type(x), np.number)) else (str(x) if x != "" else "")
            if fmt == 'pct': return f"{float(x)*100:.2f}%" if (isinstance(x, (int, float)) and x <= 1) else (f"{float(x):.2f}%" if isinstance(x, (int, float)) else (str(x) if x != "" else ""))
            if fmt == 'dec2': return f"{float(x):.2f}" if isinstance(x, (int, float)) else (str(x) if x != "" else "")
            if fmt == 'score': return f"{float(x):.2f}" if isinstance(x, (int, float)) else (str(x) if x != "" else "")
            return str(x) if x != "" else ""
        
        def _format_column_vectorized(series, fmt_type):
            """Vectorized formatting function for pandas Series"""
            if series.empty:
                return series
            # Handle NaN and empty strings
            mask_valid = series.notna() & (series != "")
            if not mask_valid.any():
                return series.astype(str)
            
            result = series.copy().astype(str)
            
            if fmt_type == 'cur':
                numeric_mask = pd.to_numeric(series, errors='coerce').notna() & mask_valid
                if numeric_mask.any():
                    numeric_vals = pd.to_numeric(series[numeric_mask], errors='coerce')
                    result.loc[numeric_mask] = numeric_vals.apply(lambda x: f"${x:,.2f}")
            elif fmt_type == 'pct':
                numeric_mask = pd.to_numeric(series, errors='coerce').notna() & mask_valid
                if numeric_mask.any():
                    numeric_vals = pd.to_numeric(series[numeric_mask], errors='coerce')
                    # Check if values are already percentages (>1) or decimals (<=1)
                    pct_mask = numeric_vals <= 1
                    result.loc[numeric_mask & pct_mask] = numeric_vals[pct_mask].apply(lambda x: f"{x*100:.2f}%")
                    result.loc[numeric_mask & ~pct_mask] = numeric_vals[~pct_mask].apply(lambda x: f"{x:.2f}%")
            elif fmt_type in ['dec2', 'score']:
                numeric_mask = pd.to_numeric(series, errors='coerce').notna() & mask_valid
                if numeric_mask.any():
                    numeric_vals = pd.to_numeric(series[numeric_mask], errors='coerce')
                    result.loc[numeric_mask] = numeric_vals.apply(lambda x: f"{x:.2f}")
            
            # Preserve empty strings and NaN as empty strings
            result.loc[~mask_valid] = ""
            return result
        if hierarchical_results:
            show_allocation_columns = bool(allocation_calc) and not hide_allocation_factors
            display_df = build_hierarchical_display(
                hierarchical_results, var_to_node, hierarchy_data_display,
                allocation_lookup=allocation_calc if show_allocation_columns else None,
                n_adults=n_a,
                n_children=n_c
            )
            exp_cols = ['Expenditure Category', 'Reported $', 'Coefficient of Variation', 'Quality', 'Allocated $']
            alloc_cols = ['Shared %', 'Child Intensity', 'Shared $', 'Exclusive (Adult) $', 'Exclusive (Child) $']
            display_cols = exp_cols + [c for c in alloc_cols if c in display_df.columns]
            display_df = display_df[[c for c in display_cols if c in display_df.columns]].copy()
            for col in ['Reported $', 'Allocated $', 'Shared $', 'Exclusive (Adult) $', 'Exclusive (Child) $']:
                if col in display_df.columns:
                    display_df[col] = _format_column_vectorized(display_df[col], 'cur')
            if 'Coefficient of Variation' in display_df.columns:
                display_df['Coefficient of Variation'] = _format_column_vectorized(display_df['Coefficient of Variation'], 'dec2')
            if 'Shared %' in display_df.columns:
                display_df['Shared %'] = _format_column_vectorized(display_df['Shared %'], 'pct')
            if 'Child Intensity' in display_df.columns:
                display_df['Child Intensity'] = _format_column_vectorized(display_df['Child Intensity'], 'score')
            st.dataframe(display_df, use_container_width=True, height=400, hide_index=True)
        else:
            results_df_filtered = results_df
            if var_to_node and 'Spending Code' in results_df.columns:
                # Vectorized filtering: pre-compute levels for all codes
                codes = results_df['Spending Code'].values
                levels = [_level_from_node(var_to_node, code) for code in codes]
                keep_mask = pd.Series([(level is None or level <= max_granularity_level) for level in levels], index=results_df.index)
                results_df_filtered = results_df[keep_mask]
            need = ['Spending Description', 'Mean Dollars Per Year', 'Coefficient of Variation', 'Data Quality Category']
            fallback_df = results_df_filtered[[c for c in need if c in results_df_filtered.columns]].copy()
            fallback_df = fallback_df.rename(columns={
                'Spending Description': 'Expenditure Category',
                'Mean Dollars Per Year': 'Reported $',
                'Coefficient of Variation': 'Coefficient of Variation',
                'Data Quality Category': 'Quality'
            })
            fallback_df['Allocated $'] = ""
            if 'Quality' in fallback_df.columns:
                f_mask = fallback_df['Quality'].fillna('F').astype(str).str.upper().str.strip() == 'F'
                fallback_df.loc[f_mask, 'Reported $'] = ""
                fallback_df.loc[f_mask, 'Allocated $'] = ""
            if allocation_display and not hide_allocation_factors:
                for c in ['Shared %', 'Child Intensity', 'Shared $', 'Exclusive (Child) $', 'Exclusive (Adult) $']:
                    fallback_df[c] = ""
                if 'Quality' in fallback_df.columns:
                    for c in ['Shared %', 'Child Intensity', 'Shared $', 'Exclusive (Child) $', 'Exclusive (Adult) $']:
                        fallback_df.loc[f_mask, c] = ""
            exp_cols = ['Expenditure Category', 'Reported $', 'Coefficient of Variation', 'Quality', 'Allocated $']
            fallback_df = fallback_df[[c for c in exp_cols + (['Shared %', 'Child Intensity', 'Shared $', 'Exclusive (Adult) $', 'Exclusive (Child) $'] if allocation_display and not hide_allocation_factors else []) if c in fallback_df.columns]].copy()
            for col in ['Reported $', 'Allocated $', 'Shared $', 'Exclusive (Adult) $', 'Exclusive (Child) $']:
                if col in fallback_df.columns:
                    fallback_df[col] = _format_column_vectorized(fallback_df[col], 'cur')
            if 'Coefficient of Variation' in fallback_df.columns:
                fallback_df['Coefficient of Variation'] = _format_column_vectorized(fallback_df['Coefficient of Variation'], 'dec2')
            if 'Shared %' in fallback_df.columns:
                fallback_df['Shared %'] = _format_column_vectorized(fallback_df['Shared %'], 'pct')
            if 'Child Intensity' in fallback_df.columns:
                fallback_df['Child Intensity'] = _format_column_vectorized(fallback_df['Child Intensity'], 'score')
            st.dataframe(fallback_df, use_container_width=True, height=400, hide_index=True)
        with st.expander("Quality and Child Intensity legends"):
            st.markdown(_quality_help + "  \n" + _child_help)
        
        # Export options - Single download button
        st.subheader("📥 Export Results")
        
        try:
            from io import BytesIO
            from openpyxl import load_workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            
            output = BytesIO()
            
            # Get all available filter variables and their options (use config for column names)
            _filter_label_keys = ['PROV', 'HHTYPE6', 'HHSIZE', 'DWELTYP', 'TENURE', 'RP_AGEGRP', 'RP_GENDER', 'RP_MARSTAT', 'RP_EDUC', 'SP_AGEGRP', 'SP_EDUC', 'P0TO4YN', 'P5TO15YN', 'VEHICLEYN', 'HH_MAJINCSRC']
            all_filter_vars = {k: get_unique_values(df, YEAR_CONFIG.get_filter_column(k)) for k in _filter_label_keys}
            var_name_map = {k: YEAR_CONFIG.get_filter_column(k) for k in _filter_label_keys}
            
            # Create single sheet with all sections
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # Create empty dataframe to start
                all_data = []
                
                # TOP SECTION: Source and Filters
                all_data.append([YEAR_CONFIG.get_excel_title()])
                all_data.append([""])
                all_data.append(["Source:"])
                all_data.append([f"Statistics Canada. Survey of Household Spending, {year_choice}. " +
                                "Public Use Microdata File. Statistics Canada Catalogue no. 62M0004X. " +
                                "This does not constitute an endorsement by Statistics Canada of this product."])
                all_data.append([""])
                all_data.append(["Generated:", pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")])
                all_data.append([""])
                all_data.append(["Filter Criteria:"])
                all_data.append(["Variable", "Selected Value", "All Available Options"])
                
                # Add filter information
                filter_labels = {
                    'PROV': 'Province',
                    'HHTYPE6': 'Household type',
                    'HHSIZE': 'Household size',
                    'DWELTYP': 'Type of dwelling',
                    'TENURE': 'Dwelling tenure',
                    'RP_AGEGRP': 'Reference person - Age group',
                    'RP_GENDER': 'Reference person - Gender',
                    'RP_MARSTAT': 'Reference person - Marital status',
                    'RP_EDUC': 'Reference person - Education',
                    'SP_AGEGRP': 'Spouse - Age group',
                    'SP_EDUC': 'Spouse - Education',
                    'P0TO4YN': 'Presence of persons aged 0 to 4 years',
                    'P5TO15YN': 'Presence of persons aged 5 to 15 years',
                    'VEHICLEYN': 'Owned, leased or operated a vehicle',
                    'HH_MAJINCSRC': 'Household - Major source of income'
                }
                
                for var, label in filter_labels.items():
                    if var in all_filter_vars and all_filter_vars[var]:
                        actual_var = var_name_map.get(var, var)
                        selected_val = st.session_state.filters.get(actual_var, None)
                        if selected_val is not None:
                            # Handle both single values and lists
                            if isinstance(selected_val, list):
                                if len(selected_val) > 0:
                                    selected_labels = []
                                    for val in selected_val:
                                        lbl = format_value(var, val)
                                        selected_labels.append(f"{lbl} ({val})")
                                    selected_display = "; ".join(selected_labels)
                                else:
                                    selected_display = "All"
                            else:
                                selected_label = format_value(var, selected_val)
                                selected_display = f"{selected_label} ({selected_val})"
                        else:
                            selected_display = "All"
                        
                        # Get all available options
                        options_list = []
                        for val in sorted(all_filter_vars[var]):
                            opt_label = format_value(var, val)
                            options_list.append(f"{opt_label} ({val})")
                        options_str = "; ".join(options_list[:10])  # Limit to first 10 for display
                        if len(options_list) > 10:
                            options_str += f"; ... ({len(options_list)} total options)"
                        
                        all_data.append([label, selected_display, options_str])
                
                # Add income range filter if applied
                if st.session_state.get('income_range') is not None:
                    income_range = st.session_state.income_range
                    all_data.append(["Household Total Income Range:", f"${income_range[0]:,.0f} to ${income_range[1]:,.0f}"])
                
                all_data.append([""])
                all_data.append(["Number of Records Matching Criteria:", st.session_state.get('filtered_count', 'N/A')])
                
                all_data.append([""])
                all_data.append([""])
                
                # Prepare allocation and hierarchy for middle section and expenditure table
                allocation_export = st.session_state.get('allocation_input')
                force_shared_allocation = st.session_state.get("force_shared_allocation", False)
                hide_allocation_factors = st.session_state.get("hide_allocation_factors", False)
                use_lower_level_weights = st.session_state.get("use_lower_level_weights", False) and not hide_allocation_factors
                hierarchy_data_export = st.session_state.get('hierarchy_data', hierarchy_data)
                hierarchical_results_export_full, var_to_node_export = organize_hierarchical_results(results_df, hierarchy_data_export)
                granularity_level = int(st.session_state.get("granularity_level", 7))
                max_granularity_level = granularity_level - 1
                
                # Get results at Level 7 for lower-level weight calculation
                hierarchical_results_export_level7 = filter_results_by_granularity(
                    hierarchical_results_export_full,
                    var_to_node_export,
                    6  # Level 7 is max level (0-based: 6)
                ) if hierarchical_results_export_full else []
                
                # Get results at requested granularity level
                hierarchical_results_export = filter_results_by_granularity(
                    hierarchical_results_export_full,
                    var_to_node_export,
                    max_granularity_level
                )
                
                gran_alloc = compute_granular_allocation(hierarchical_results_export, var_to_node_export, hierarchy_data_export) if hierarchical_results_export else {}
                
                # If using lower-level weights, compute aggregated allocation factors from Level 7
                # IMPORTANT: When switch is OFF, use ONLY original spreadsheet values
                # When switch is ON, use ONLY aggregated values (completely replace)
                if use_lower_level_weights and hierarchical_results_export_level7 and hierarchical_results_export:
                    # Get original allocation for Level 7 calculations
                    original_allocation_export_calc = _force_shared_allocation(allocation_export) if force_shared_allocation else allocation_export
                    
                    # Compute aggregated allocation factors from Level 7 up to requested level
                    aggregated_allocation_export = compute_aggregated_allocation_factors(
                        hierarchical_results_export_level7,
                        hierarchical_results_export,
                        var_to_node_export,
                        hierarchy_data_export,
                        int(st.session_state.get('allocation_n_adults', 2)),
                        int(st.session_state.get('allocation_n_children', 0)),
                        original_allocation_export_calc if original_allocation_export_calc else {}
                    )
                    
                    # When switch is ON, use ONLY aggregated values (completely replace, don't merge)
                    # Only include nodes that have aggregated values
                    allocation_export_calc = {}
                    for item in hierarchical_results_export:
                        var_code = item['var_code'] if isinstance(item, dict) else item
                        if var_code in aggregated_allocation_export:
                            allocation_export_calc[var_code] = aggregated_allocation_export[var_code].copy()
                else:
                    # When switch is OFF, use ONLY original spreadsheet values
                    allocation_export_calc = _force_shared_allocation(allocation_export) if force_shared_allocation else allocation_export
                
                # Total Consumption and Gifts = TC001 + MG001
                total_consumption_gifts = 0
                if 'Spending Code' in results_df.columns and 'Mean Dollars Per Year' in results_df.columns:
                    total_consumption_gifts = (
                        results_df.loc[results_df['Spending Code'] == 'TC001', 'Mean Dollars Per Year'].sum() +
                        results_df.loc[results_df['Spending Code'] == 'MG001', 'Mean Dollars Per Year'].sum()
                    )
                n_a = int(st.session_state.get('allocation_n_adults', 2))
                n_c = int(st.session_state.get('allocation_n_children', 0))
                
                # Allocation totals across granular allocation rows (for middle-section percentages and dollars)
                total_shared_d = 0.0
                total_excl_adult_d = 0.0
                total_excl_child_d = 0.0
                if allocation_export_calc and hierarchical_results_export:
                    for item in hierarchical_results_export:
                        ga = gran_alloc.get(item['var_code'], np.nan)
                        if not _has_granular_value(ga):
                            continue
                        lookup = allocation_export_calc.get(item['var_code'], {})
                        v1, v2 = lookup.get('shared_pct'), lookup.get('child_intensity')
                        shared, excl_c, excl_a = _allocation_split(item['mean'], v1, v2, n_a, n_c)
                        total_shared_d += shared
                        total_excl_adult_d += n_a * excl_a
                        total_excl_child_d += n_c * excl_c
                total_alloc = total_shared_d + total_excl_adult_d + total_excl_child_d
                pct_shared = (total_shared_d / total_alloc * 100) if total_alloc else 0
                # Per-adult and per-child amounts and percentages (not aggregates)
                per_adult_d = (total_excl_adult_d / n_a) if n_a else 0
                per_child_d = (total_excl_child_d / n_c) if n_c else 0
                pct_per_adult = (per_adult_d / total_alloc * 100) if total_alloc else 0
                pct_per_child = (per_child_d / total_alloc * 100) if total_alloc else 0
                
                # Middle section: Total Consumption and Gifts, N Adults/Children, then header "Dollars"|"Percent" and 3 rows: Shared Spending, Exclusive per Adult, Exclusive per Child
                all_data.append(["Total Consumption and Gifts", round(total_consumption_gifts, 0) if total_consumption_gifts else 0])
                all_data.append(["Number of Adults", int(n_a)])
                all_data.append(["Number of Children", int(n_c)])
                if not hide_allocation_factors:
                    all_data.append(["", "Dollars", "Percent"])  # header for the 3 allocation rows
                    all_data.append(["Shared Spending", round(total_shared_d, 0) if allocation_export_calc else "", (round(pct_shared / 100, 4) if allocation_export_calc and total_alloc else "")])
                    all_data.append(["Exclusive Spending per Adult", round(per_adult_d, 0) if allocation_export_calc else "", (round(pct_per_adult / 100, 4) if allocation_export_calc and total_alloc else "")])
                    all_data.append(["Exclusive Spending per Child", round(per_child_d, 0) if allocation_export_calc else "", (round(pct_per_child / 100, 4) if allocation_export_calc and total_alloc else "")])
                
                all_data.append([""])
                all_data.append([""])
                
                # BOTTOM SECTION: Expenditure Categories (new column titles, no CV, suppress F)
                all_data.append(["Results"])
                exp_header = ["Expenditure Category", "Reported $", "Coefficient of Variation", "Quality", "Allocated $"]
                if allocation_export_calc is not None and not hide_allocation_factors:
                    exp_header += ["Shared %", "Child Intensity", "Shared $", "Exclusive (Adult) $", "Exclusive (Child) $"]
                all_data.append(exp_header)
                exp_header_excel_row = len(all_data)
                exp_num_cols = len(exp_header)
                
                if hierarchical_results_export:
                    INDENT_PER_LEVEL = 2
                    for item in hierarchical_results_export:
                        quality = str(item.get('quality', 'F') or 'F').strip().upper()
                        is_f = (quality == 'F')
                        level = int(item['level']) if item.get('level') is not None else 0
                        indent = " " * (level * INDENT_PER_LEVEL)
                        var_code = item['var_code']
                        description = item['description']
                        ga = gran_alloc.get(var_code, np.nan)
                        ga_display = "" if is_f else ("" if (ga is None or (isinstance(ga, float) and np.isnan(ga))) else round(ga, 2))
                        row = [
                            f"{indent}{description}",
                            "" if is_f else round(item['mean'], 2),
                            round(item.get('cv'), 2) if item.get('cv') is not None and not pd.isna(item.get('cv')) else "",
                            quality,
                            ga_display
                        ]
                        if allocation_export_calc is not None and not hide_allocation_factors:
                            if is_f:
                                row.extend(["", "", "", "", ""])
                            else:
                                show_alloc = _has_granular_value(ga)
                                lookup = allocation_export_calc.get(var_code, {}) if show_alloc else {}
                                v1 = lookup.get('shared_pct')
                                v2 = lookup.get('child_intensity')
                                s = (v1 / 100 if (isinstance(v1, (int, float)) and v1 > 1) else v1) if v1 is not None else None
                                row.append(round(s, 4) if s is not None else "")  # fraction for Excel 0.00%
                                row.append(round(v2, 2) if v2 is not None and isinstance(v2, (int, float)) else "")
                                if show_alloc:
                                    shared, excl_c, excl_a = _allocation_split(item['mean'], v1, v2, n_a, n_c)
                                    row.append(round(shared, 2))
                                    row.append(round(excl_a, 2))
                                    row.append(round(excl_c, 2))
                                else:
                                    row.extend(["", "", ""])
                        all_data.append(row)
                else:
                    need = ['Spending Description', 'Mean Dollars Per Year', 'Coefficient of Variation', 'Data Quality Category']
                    results_export = results_df[[c for c in need if c in results_df.columns]].copy()
                    # Vectorized processing: compute quality flags and F masks first
                    if 'Data Quality Category' in results_export.columns:
                        quals = results_export['Data Quality Category'].fillna('F').astype(str).str.strip().str.upper()
                        is_f_mask = (quals == 'F')
                    else:
                        quals = pd.Series(['F'] * len(results_export))
                        is_f_mask = pd.Series([True] * len(results_export))
                    
                    # Build rows vectorized
                    for idx in results_export.index:
                        qual = quals.loc[idx]
                        is_f = is_f_mask.loc[idx]
                        mean_val = results_export.loc[idx, 'Mean Dollars Per Year'] if 'Mean Dollars Per Year' in results_export.columns else None
                        cv_val = results_export.loc[idx, 'Coefficient of Variation'] if 'Coefficient of Variation' in results_export.columns else None
                        desc = results_export.loc[idx, 'Spending Description'] if 'Spending Description' in results_export.columns else ''
                        
                        data_row = [
                            desc,
                            "" if is_f else (round(mean_val, 2) if mean_val is not None and not pd.isna(mean_val) else ""),
                            round(cv_val, 2) if cv_val is not None and not pd.isna(cv_val) else "",
                            qual,
                            "" if is_f else ""
                        ]
                        if allocation_export_calc is not None and not hide_allocation_factors:
                            data_row.extend(["", "", "", "", ""])
                        all_data.append(data_row)
                
                # Convert to DataFrame and write
                export_df = pd.DataFrame(all_data)
                export_df.to_excel(writer, sheet_name='Spending Estimates', index=False, header=False)
            
            # Format the Excel file
            output.seek(0)
            wb = load_workbook(output)
            ws = wb['Spending Estimates']
            
            # Set print area and page setup
            max_row = ws.max_row
            max_col = ws.max_column
            ws.print_area = '$A$1:$J$339'
            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
            
            # Style headers and important rows
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            title_font = Font(bold=True, size=12)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Format title row
            ws['A1'].font = title_font
            ws.merge_cells(f'A1:{get_column_letter(max_col)}1')
            
            # Format section headers
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_row), 1):
                cell_value = str(row[0].value) if row[0].value else ""
                
                # Format section headers
                is_header = any(keyword in cell_value for keyword in ["Source:", "Filter Criteria:", 
                                                             "Results", "By Expenditure Category", "Spending Category Breakdown", "Individual Spending Code Breakdown", "TOTAL"])
                if is_header:
                    for cell in row:
                        cell.font = Font(bold=True, size=11)
                        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            
            # Column widths: A=100, B=17.1, C=13.5; rest 10.5. A left-justified, B and C right-justified.
            ws.column_dimensions['A'].width = 100.0
            ws.column_dimensions['B'].width = 17.1
            ws.column_dimensions['C'].width = 13.5
            for c in range(4, exp_num_cols + 1):
                ws.column_dimensions[get_column_letter(c)].width = 10.5
            align_left = Alignment(horizontal='left', wrap_text=True, vertical='top')
            align_left_no_wrap = Alignment(horizontal='left', wrap_text=False, vertical='top')
            wrap_right = Alignment(horizontal='right', wrap_text=True, vertical='top')
            wrap_right_no_wrap = Alignment(horizontal='right', wrap_text=False, vertical='top')
            for r in range(1, max_row + 1):
                wrap_off = (6 <= r <= 37)
                al1 = align_left_no_wrap if wrap_off else align_left
                wr2 = wrap_right_no_wrap if wrap_off else wrap_right
                wr3 = wrap_right_no_wrap if wrap_off else wrap_right
                ws.cell(row=r, column=1).alignment = al1
                if max_col >= 2:
                    ws.cell(row=r, column=2).alignment = wr2
                if max_col >= 3:
                    c3_align = align_left_no_wrap if (9 <= r <= 23) else wr3
                    ws.cell(row=r, column=3).alignment = c3_align
            for c in range(4, exp_num_cols + 1):
                ws.cell(row=exp_header_excel_row, column=c).alignment = wrap_right
            # Rows 6–37: no wrap
            for r in range(6, min(38, max_row + 1)):
                for c in range(1, max_col + 1):
                    cur = ws.cell(row=r, column=c).alignment
                    ws.cell(row=r, column=c).alignment = Alignment(horizontal=cur.horizontal, wrap_text=False, vertical=cur.vertical)
            
            # Expenditure: 1=Expenditure Category (text), 2=Reported $ (currency), 3=Coefficient of Variation (0.00),
            # 4=Quality (text), 5=Allocated $ (currency), 6=Shared % (0.00%), 7=Child Intensity (#.00),
            # 8=Shared $, 9=Exclusive (Adult) $, 10=Exclusive (Child) $ (currency)
            fmt_currency2 = '$#,##0.00'
            fmt_currency0 = '$#,##0'
            fmt_pct = '0.00%'
            fmt_dec2 = '#.00'
            for r in range(exp_header_excel_row + 1, max_row + 1):
                for c in range(1, exp_num_cols + 1):
                    cell = ws.cell(row=r, column=c)
                    if c == 1 or c == 4:
                        cell.number_format = '@'
                    elif c == 2 or c == 5:
                        cell.number_format = fmt_currency2
                    elif c == 3 and exp_num_cols >= 3:
                        cell.number_format = fmt_dec2
                    elif c == 6 and exp_num_cols >= 6:
                        cell.number_format = fmt_pct
                    elif c == 7 and exp_num_cols >= 7:
                        cell.number_format = fmt_dec2
                    elif c in (8, 9, 10) and exp_num_cols >= c:
                        cell.number_format = fmt_currency2
            
            # Middle section: Total Consumption and Gifts, N Adults/Children, header Dollars|Percent, 3 allocation rows (B=$, C=%)
            for r in range(1, max_row + 1):
                a1 = ws.cell(row=r, column=1).value
                if a1 and str(a1).strip() == "Total Consumption and Gifts":
                    ws.cell(row=r, column=2).number_format = fmt_currency0
                    ws.cell(row=r + 1, column=2).number_format = '0'
                    ws.cell(row=r + 2, column=2).number_format = '0'
                    if not hide_allocation_factors:
                        # r+3: header "Dollars"|"Percent" (text)
                        ws.cell(row=r + 4, column=2).number_format = fmt_currency0
                        ws.cell(row=r + 4, column=3).number_format = fmt_pct
                        ws.cell(row=r + 5, column=2).number_format = fmt_currency0
                        ws.cell(row=r + 5, column=3).number_format = fmt_pct
                        ws.cell(row=r + 6, column=2).number_format = fmt_currency0
                        ws.cell(row=r + 6, column=3).number_format = fmt_pct
                    break
            
            # Auto-adjust column widths for columns beyond exp_num_cols (A=100, B=17.1, C=13.5, 4..exp=10.5 already set)
            for col in ws.columns:
                col_letter = get_column_letter(col[0].column)
                col_idx = col[0].column
                if col_idx <= exp_num_cols:
                    continue
                max_length = 0
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except Exception:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[col_letter].width = adjusted_width
            
            # Save formatted workbook
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            excel_data = output.read()
            
            col_left, col_right = st.columns([1, 3])
            with col_left:
                st.download_button(
                    label="Download All Results (Excel)",
                    data=excel_data,
                    file_name="spending_estimates_all.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="excel_all",
                    type="primary",
                    use_container_width=True
                )
        except ImportError:
            st.warning("Excel export requires openpyxl and Pillow. Install with: pip install openpyxl pillow")
        except Exception as e:
            st.error(f"Error creating Excel file: {e}")
            import traceback
            st.text(traceback.format_exc())
    
if __name__ == "__main__":
    main()
